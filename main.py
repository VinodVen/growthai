import os
import re
import bcrypt
import stripe
import stripe.checkout
import smtplib
import hmac
import hashlib
import base64
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify
from werkzeug.middleware.proxy_fix import ProxyFix
from dotenv import load_dotenv
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

load_dotenv()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

ENV = os.getenv("FLASK_ENV", "development")
DEBUG = os.getenv("DEBUG", "False").lower() == "true"
IS_PRODUCTION = ENV == "production"

app.config["ENV"] = ENV
app.config["DEBUG"] = DEBUG
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key-12345")
app.config["SESSION_COOKIE_SECURE"] = IS_PRODUCTION
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = 60 * 60 * 24 * 30  # 30 days

# Stripe
stripe.api_key = os.getenv("STRIPE_SECRET_KEY", "sk_test_fake")
STRIPE_PRICE_STARTER = os.getenv("STRIPE_PRICE_ID_STARTER") or os.getenv("STRIPE_PRICE_ID")
STRIPE_PRICE_PRO = os.getenv("STRIPE_PRICE_ID_PRO") or os.getenv("STRIPE_PRICE_ID")

# OpenAI — use requests directly to avoid httpx/proxies version conflicts
import requests as _requests
openai_init_error = None
_openai_api_key = os.getenv("OPENAI_API_KEY", "").strip()
client = True if _openai_api_key else None  # just a flag
if not _openai_api_key:
    openai_init_error = "OPENAI_API_KEY not set"
    print("Warning: OPENAI_API_KEY not set.")

def _call_openai(prompt, max_tokens=150):
    """Call OpenAI API directly via requests, avoiding httpx version conflicts."""
    response = _requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {_openai_api_key}",
            "Content-Type": "application/json"
        },
        json={
            "model": "gpt-4o-mini",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens,
            "temperature": 0.7
        },
        timeout=15
    )
    data = response.json()
    if "error" in data:
        raise Exception(data["error"]["message"])
    return data["choices"][0]["message"]["content"]

# Twilio SMS
try:
    from twilio.rest import Client as TwilioClient
    twilio_sid = os.getenv("TWILIO_ACCOUNT_SID")
    twilio_token = os.getenv("TWILIO_AUTH_TOKEN")
    twilio_phone = os.getenv("TWILIO_PHONE_NUMBER")
    if twilio_sid and twilio_token:
        twilio_client = TwilioClient(twilio_sid, twilio_token)
    else:
        twilio_client = None
except Exception as e:
    twilio_client = None
    twilio_phone = None
    print(f"Warning: Twilio not configured: {e}")

# Plan limits
PLAN_LIMITS = {
    "free":    {"customers": 50,  "campaigns_month": 10},
    "starter": {"customers": 500, "campaigns_month": None},
    "pro":     {"customers": None,"campaigns_month": None},
}

CAMPAIGN_TYPES = {
    "come_back":   "Come Back Offer",
    "weekend":     "Weekend Special",
    "lunch":       "Lunch Deal",
    "dinner":      "Dinner Special",
    "birthday":    "Birthday Special",
    "loyalty":     "Loyalty Reward",
    "happy_hour":  "Happy Hour",
    "new_item":    "New Item Launch",
    "promotion":   "General Promotion",
    "invitation":  "Business Invitation",
    "custom":      "Custom Message",
}

# Database
db_url = os.getenv("DATABASE_URL", "sqlite:///restaurant.db")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
if not db_url.startswith("sqlite"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_size": 10,
        "pool_recycle": 3600,
        "pool_pre_ping": True,
    }

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# ============================================
# MODELS
# ============================================

def make_slug(name, uid):
    """Generate a URL-safe slug from business name + id."""
    base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return f"{base}-{uid}"

class Business(db.Model):
    __tablename__ = "businesses"
    id = db.Column(db.Integer, primary_key=True)
    business_name = db.Column(db.String(200), nullable=False)
    owner_name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(200), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    plan = db.Column(db.String(50), default="free")
    stripe_customer_id = db.Column(db.String(200))
    address = db.Column(db.String(300))
    phone = db.Column(db.String(50))
    website = db.Column(db.String(200))
    slug = db.Column(db.String(200), unique=True)          # e.g. marios-pizza-3
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class BusinessProfile(db.Model):
    """Stores automation settings filled in by the business owner."""
    __tablename__ = "business_profiles"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), unique=True, nullable=False)
    cuisine_type = db.Column(db.String(100))               # Italian, Mexican, Indian…
    signature_dish = db.Column(db.String(200))             # "Wood-fired pizza"
    special_offer = db.Column(db.String(200))              # "15% off", "free dessert"
    slow_days = db.Column(db.String(100))                  # "Mon,Tue"
    peak_days = db.Column(db.String(100))                  # "Fri,Sat,Sun"
    tone = db.Column(db.String(50), default="friendly")    # friendly|professional|fun
    timezone = db.Column(db.String(50), default="America/Chicago")
    auto_welcome = db.Column(db.Boolean, default=True)     # welcome new customers
    auto_weekly = db.Column(db.Boolean, default=False)     # weekly special every week
    auto_flash = db.Column(db.Boolean, default=False)      # flash deal on slow days
    auto_birthday = db.Column(db.Boolean, default=True)
    auto_winback = db.Column(db.Boolean, default=True)
    weekly_send_day = db.Column(db.String(20), default="Tuesday")
    setup_complete = db.Column(db.Boolean, default=False)
    last_weekly_summary = db.Column(db.DateTime, nullable=True)  # track when summary last sent
    auto_review = db.Column(db.Boolean, default=True)         # send review request
    review_days = db.Column(db.Integer, default=3)            # days after signup to send review
    google_review_url = db.Column(db.String(300))             # google review link
    auto_loyalty = db.Column(db.Boolean, default=True)        # loyalty milestone emails
    loyalty_reward_visits = db.Column(db.Integer, default=5)  # visits to earn a reward
    language = db.Column(db.String(20), default="English")    # message language for AI
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Customer(db.Model):
    __tablename__ = "customers"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    first_name = db.Column(db.String(120), nullable=False)
    last_name = db.Column(db.String(120))
    email = db.Column(db.String(200), nullable=True)
    phone = db.Column(db.String(50))
    dob = db.Column(db.String(50))
    notes = db.Column(db.Text)
    tags = db.Column(db.String(300))  # comma-separated: VIP,Regular,Corporate
    visit_count = db.Column(db.Integer, default=0)
    last_visit = db.Column(db.DateTime, nullable=True)        # last time customer visited
    loyalty_points = db.Column(db.Integer, default=0)         # cumulative loyalty points
    unsubscribed = db.Column(db.Boolean, default=False)
    sms_opted_in = db.Column(db.Boolean, default=False)    # TCPA compliance
    sms_opted_in_at = db.Column(db.DateTime, nullable=True)
    sms_opt_in_ip = db.Column(db.String(60), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Campaign(db.Model):
    __tablename__ = "campaigns"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    customer_name = db.Column(db.String(200), nullable=False)
    customer_email = db.Column(db.String(200), nullable=True)
    customer_phone = db.Column(db.String(50))
    campaign_type = db.Column(db.String(50), nullable=False)
    message = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(50), default="draft")
    scheduled_at = db.Column(db.DateTime, nullable=True)
    open_count = db.Column(db.Integer, default=0)
    opened_at = db.Column(db.DateTime, nullable=True)
    click_count = db.Column(db.Integer, default=0)
    clicked_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ContactMessage(db.Model):
    __tablename__ = "contact_messages"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Promotion(db.Model):
    __tablename__ = "promotions"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.String(50))
    category = db.Column(db.String(50), default="promotion")
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AutomationRule(db.Model):
    __tablename__ = "automation_rules"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    rule_type = db.Column(db.String(50), nullable=False)  # birthday, reengagement, welcome
    active = db.Column(db.Boolean, default=False)
    message_template = db.Column(db.Text)
    days_threshold = db.Column(db.Integer, default=30)  # days before birthday or days since contact
    last_run = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Journey(db.Model):
    __tablename__ = "journeys"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    trigger = db.Column(db.String(50), default="signup")  # signup, birthday, winback, manual
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # AJO-style controls
    allow_reentry = db.Column(db.Boolean, default=False)       # re-enroll after completion
    freq_cap_per_week = db.Column(db.Integer, default=0)       # 0 = no cap
    goal = db.Column(db.String(50), default="none")            # none, visit, email_open
    journey_steps = db.relationship("JourneyStep", backref="journey", lazy=True, order_by="JourneyStep.step_order")

class JourneyStep(db.Model):
    __tablename__ = "journey_steps"
    id = db.Column(db.Integer, primary_key=True)
    journey_id = db.Column(db.Integer, db.ForeignKey("journeys.id"), nullable=False)
    step_order = db.Column(db.Integer, default=0)
    delay_days = db.Column(db.Integer, default=0)
    message_type = db.Column(db.String(50), default="promotion")
    message_text = db.Column(db.Text)
    use_ai = db.Column(db.Boolean, default=True)
    channel = db.Column(db.String(20), default="email")  # email, sms, both
    condition = db.Column(db.String(50), default="none")  # none, visited, not_visited
    sms_fallback = db.Column(db.Boolean, default=False)      # send SMS if email not opened
    sms_fallback_days = db.Column(db.Integer, default=2)     # days to wait before checking

class CustomerJourney(db.Model):
    __tablename__ = "customer_journeys"
    id = db.Column(db.Integer, primary_key=True)
    journey_id = db.Column(db.Integer, db.ForeignKey("journeys.id"), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    next_step_order = db.Column(db.Integer, default=0)
    enrolled_at = db.Column(db.DateTime, default=datetime.utcnow)
    next_step_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed = db.Column(db.Boolean, default=False)
    active = db.Column(db.Boolean, default=True)
    last_campaign_id = db.Column(db.Integer, db.ForeignKey("campaigns.id"), nullable=True)
    waiting_for_open = db.Column(db.Boolean, default=False)  # True = sent email, checking open
    open_check_at = db.Column(db.DateTime, nullable=True)    # when to check if email was opened

class Audience(db.Model):
    __tablename__ = "audiences"
    id = db.Column(db.Integer, primary_key=True)
    business_id = db.Column(db.Integer, db.ForeignKey("businesses.id"), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    rules = db.Column(db.Text, default="[]")  # JSON array of rule objects
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class CampaignTypeModel(db.Model):
    __tablename__ = "campaign_type_options"
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    label = db.Column(db.String(100), nullable=False)
    active = db.Column(db.Boolean, default=True)
    sort_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

def _run_migrations():
    """Run DB migrations in background so startup doesn't block gunicorn."""
    import time
    time.sleep(2)  # let app fully start first
    with app.app_context():
        _do_migrations()

def _do_migrations():
    from sqlalchemy import text
    migrations = [
        "ALTER TABLE businesses ADD COLUMN IF NOT EXISTS address VARCHAR(300)",
        "ALTER TABLE businesses ADD COLUMN IF NOT EXISTS phone VARCHAR(50)",
        "ALTER TABLE businesses ADD COLUMN IF NOT EXISTS website VARCHAR(200)",
        "ALTER TABLE businesses ADD COLUMN IF NOT EXISTS slug VARCHAR(200)",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS unsubscribed BOOLEAN DEFAULT FALSE",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS sms_opted_in BOOLEAN DEFAULT FALSE",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS sms_opted_in_at TIMESTAMP",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS sms_opt_in_ip VARCHAR(60)",
        "ALTER TABLE campaigns ADD COLUMN IF NOT EXISTS scheduled_at TIMESTAMP",
        "ALTER TABLE campaigns ADD COLUMN IF NOT EXISTS open_count INTEGER DEFAULT 0",
        "ALTER TABLE campaigns ADD COLUMN IF NOT EXISTS opened_at TIMESTAMP",
        "ALTER TABLE campaigns ADD COLUMN IF NOT EXISTS click_count INTEGER DEFAULT 0",
        "ALTER TABLE campaigns ADD COLUMN IF NOT EXISTS clicked_at TIMESTAMP",
        "ALTER TABLE customers ALTER COLUMN email DROP NOT NULL",
        "ALTER TABLE campaigns ALTER COLUMN customer_email DROP NOT NULL",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS notes TEXT",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS tags VARCHAR(300)",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS visit_count INTEGER DEFAULT 0",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS last_weekly_summary TIMESTAMP",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS last_visit TIMESTAMP",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS loyalty_points INTEGER DEFAULT 0",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS auto_review BOOLEAN DEFAULT TRUE",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS review_days INTEGER DEFAULT 3",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS google_review_url VARCHAR(300)",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS auto_loyalty BOOLEAN DEFAULT TRUE",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS loyalty_reward_visits INTEGER DEFAULT 5",
        "ALTER TABLE business_profiles ADD COLUMN IF NOT EXISTS language VARCHAR(20) DEFAULT 'English'",
        "CREATE TABLE IF NOT EXISTS journeys (id SERIAL PRIMARY KEY, business_id INTEGER REFERENCES businesses(id), name VARCHAR(200), description TEXT, trigger VARCHAR(50) DEFAULT 'signup', active BOOLEAN DEFAULT TRUE, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS journey_steps (id SERIAL PRIMARY KEY, journey_id INTEGER REFERENCES journeys(id), step_order INTEGER DEFAULT 0, delay_days INTEGER DEFAULT 0, message_type VARCHAR(50) DEFAULT 'promotion', message_text TEXT, use_ai BOOLEAN DEFAULT TRUE, channel VARCHAR(20) DEFAULT 'email', condition VARCHAR(50) DEFAULT 'none')",
        "CREATE TABLE IF NOT EXISTS customer_journeys (id SERIAL PRIMARY KEY, journey_id INTEGER REFERENCES journeys(id), customer_id INTEGER REFERENCES customers(id), business_id INTEGER REFERENCES businesses(id), next_step_order INTEGER DEFAULT 0, enrolled_at TIMESTAMP DEFAULT NOW(), next_step_at TIMESTAMP DEFAULT NOW(), completed BOOLEAN DEFAULT FALSE, active BOOLEAN DEFAULT TRUE)",
        "ALTER TABLE journey_steps ADD COLUMN IF NOT EXISTS sms_fallback BOOLEAN DEFAULT FALSE",
        "ALTER TABLE journey_steps ADD COLUMN IF NOT EXISTS sms_fallback_days INTEGER DEFAULT 2",
        "ALTER TABLE customer_journeys ADD COLUMN IF NOT EXISTS last_campaign_id INTEGER REFERENCES campaigns(id)",
        "ALTER TABLE customer_journeys ADD COLUMN IF NOT EXISTS waiting_for_open BOOLEAN DEFAULT FALSE",
        "ALTER TABLE customer_journeys ADD COLUMN IF NOT EXISTS open_check_at TIMESTAMP",
        "CREATE TABLE IF NOT EXISTS audiences (id SERIAL PRIMARY KEY, business_id INTEGER REFERENCES businesses(id), name VARCHAR(200), description TEXT, rules TEXT DEFAULT '[]', created_at TIMESTAMP DEFAULT NOW())",
        "ALTER TABLE journeys ADD COLUMN IF NOT EXISTS allow_reentry BOOLEAN DEFAULT FALSE",
        "ALTER TABLE journeys ADD COLUMN IF NOT EXISTS freq_cap_per_week INTEGER DEFAULT 0",
        "ALTER TABLE journeys ADD COLUMN IF NOT EXISTS goal VARCHAR(50) DEFAULT 'none'",
    ]
    if not db_url.startswith("sqlite"):
        for sql in migrations:
            try:
                with db.engine.begin() as conn:
                    conn.execute(text(sql))
            except Exception as e:
                print(f"Migration skipped: {e}")
    # Seed default campaign types if table is empty
    if CampaignTypeModel.query.count() == 0:
        defaults = [
            ("come_back",     "Come Back Offer",       0),
            ("weekend",       "Weekend Special",        1),
            ("lunch",         "Lunch Deal",             2),
            ("dinner",        "Dinner Special",         3),
            ("birthday",      "Birthday Special",       4),
            ("loyalty",       "Loyalty Reward",         5),
            ("happy_hour",    "Happy Hour",             6),
            ("new_item",      "New Item Launch",        7),
            ("promotion",     "General Promotion",      8),
            ("invitation",    "Business Invitation",    9),
            ("custom",        "Custom Message",         10),
            ("hotel_stay",    "Hotel Stay Offer",       11),
            ("hotel_upgrade", "Room Upgrade Deal",      12),
            ("hotel_event",   "Event / Package Deal",   13),
            ("hotel_loyalty", "Guest Loyalty Reward",   14),
        ]
        for key, label, order in defaults:
            db.session.add(CampaignTypeModel(key=key, label=label, sort_order=order))
        db.session.commit()
    else:
        # Ensure newer campaign types exist on existing deployments
        new_types = [
            ("invitation",    "Business Invitation",    9),
            ("hotel_stay",    "Hotel Stay Offer",       11),
            ("hotel_upgrade", "Room Upgrade Deal",      12),
            ("hotel_event",   "Event / Package Deal",   13),
            ("hotel_loyalty", "Guest Loyalty Reward",   14),
        ]
        changed = False
        for key, label, order in new_types:
            if not CampaignTypeModel.query.filter_by(key=key).first():
                db.session.add(CampaignTypeModel(key=key, label=label, sort_order=order))
                changed = True
        if changed:
            db.session.commit()

with app.app_context():
    db.create_all()
    _do_migrations()

# ============================================
# HELPERS
# ============================================

@app.template_filter('from_json')
def from_json_filter(s):
    import json as _json
    try:
        return _json.loads(s or "[]")
    except Exception:
        return []

def get_campaign_types():
    types = CampaignTypeModel.query.filter_by(active=True).order_by(CampaignTypeModel.sort_order, CampaignTypeModel.label).all()
    return {t.key: t.label for t in types}

def get_segment_customers(business_id, segment):
    """Return list of Customer objects matching the given segment."""
    now = datetime.utcnow()
    base = Customer.query.filter_by(business_id=business_id, unsubscribed=False)

    if segment == "new":
        cutoff = now - timedelta(days=30)
        return base.filter(Customer.created_at >= cutoff).all()

    elif segment == "inactive":
        cutoff = now - timedelta(days=60)
        # Customers who have never been contacted, or last campaign > 60 days ago
        all_customers = base.all()
        result = []
        for c in all_customers:
            last = Campaign.query.filter_by(
                business_id=business_id, customer_email=c.email
            ).order_by(Campaign.created_at.desc()).first()
            if not last or last.created_at < cutoff:
                result.append(c)
        return result

    elif segment == "birthday_month":
        month_str = now.strftime("-%m-")
        alt_month = f"-{now.month}-"  # handles single digit months without leading zero
        all_customers = base.all()
        return [c for c in all_customers if c.dob and (
            month_str in c.dob or alt_month in c.dob
        )]

    elif segment == "vip":
        all_customers = base.all()
        result = []
        for c in all_customers:
            count = Campaign.query.filter_by(
                business_id=business_id, customer_email=c.email
            ).count()
            if count >= 3:
                result.append(c)
        return result

    elif segment == "sms_only":
        return base.filter(Customer.phone.isnot(None), Customer.phone != "",
                           (Customer.email.is_(None)) | (Customer.email == "")).all()

    else:  # "all"
        return base.all()

def current_business():
    if "user_id" not in session:
        return None
    return Business.query.get(session["user_id"])

def get_plan_limit(plan, limit_type):
    return PLAN_LIMITS.get(plan, PLAN_LIMITS["free"]).get(limit_type)

def clean_ai_text(text):
    return (text or "").replace("###", "").replace("**", "").strip()

def get_unsubscribe_token(campaign_id):
    secret = app.secret_key.encode() if isinstance(app.secret_key, str) else app.secret_key
    sig = hmac.new(secret, str(campaign_id).encode(), hashlib.sha256).hexdigest()[:20]
    token = base64.urlsafe_b64encode(f"{campaign_id}:{sig}".encode()).decode()
    return token

def verify_unsubscribe_token(token):
    try:
        decoded = base64.urlsafe_b64decode(token.encode()).decode()
        campaign_id, sig = decoded.split(":", 1)
        secret = app.secret_key.encode() if isinstance(app.secret_key, str) else app.secret_key
        expected = hmac.new(secret, campaign_id.encode(), hashlib.sha256).hexdigest()[:20]
        if hmac.compare_digest(sig, expected):
            return int(campaign_id)
    except Exception:
        pass
    return None

# Module-level throttle: track last time background tasks ran (per process)
_last_scheduled_run = {}
_last_automation_run = {}

def process_scheduled_campaigns():
    """Send any campaigns whose scheduled_at time has passed."""
    now = datetime.utcnow()
    last = _last_scheduled_run.get("t")
    if last and (now - last).total_seconds() < 300:  # throttle to once per 5 min
        return
    _last_scheduled_run["t"] = now
    try:
        pending = Campaign.query.filter(
            Campaign.status == "scheduled",
            Campaign.scheduled_at <= now
        ).all()
        for campaign in pending:
            b = Business.query.get(campaign.business_id)
            if not b:
                continue
            token = get_unsubscribe_token(campaign.id)
            unsub_url = url_for("unsubscribe", token=token, _external=True)
            pixel_url = url_for("track_open", campaign_id=campaign.id, _external=True)
            click_url = url_for("track_click", campaign_id=campaign.id, _external=True)
            subject = f"Special Offer from {b.business_name}"
            success = send_email(
                campaign.customer_email, subject, campaign.message,
                customer_name=campaign.customer_name,
                business_name=b.business_name,
                campaign_type=campaign.campaign_type,
                unsubscribe_url=unsub_url,
                business_address=b.address or "",
                business_phone=b.phone or "",
                business_website=b.website or "",
                tracking_pixel_url=pixel_url,
                click_tracking_url=click_url,
                business_reply_email=b.email
            )
            campaign.status = "sent" if success else "failed"
        if pending:
            db.session.commit()
    except Exception as e:
        print(f"Scheduler error: {e}")

def _send_auto_campaign(business, customer, campaign_type, message):
    """Helper: create + send a campaign for an automation rule."""
    if not customer.email and not customer.phone:
        return
    if customer.unsubscribed:
        return
    campaign = Campaign(
        business_id=business.id,
        customer_name=f"{customer.first_name} {customer.last_name or ''}".strip(),
        customer_email=customer.email or "",
        customer_phone=customer.phone or "",
        campaign_type=campaign_type,
        message=message,
        status="draft",
    )
    db.session.add(campaign)
    db.session.flush()
    if customer.email:
        token = get_unsubscribe_token(campaign.id)
        unsub_url = url_for("unsubscribe", token=token, _external=True)
        pixel_url = url_for("track_open", campaign_id=campaign.id, _external=True)
        click_url = url_for("track_click", campaign_id=campaign.id, _external=True)
        success = send_email(
            customer.email, f"A message from {business.business_name}", message,
            customer_name=campaign.customer_name,
            business_name=business.business_name,
            campaign_type=campaign_type,
            unsubscribe_url=unsub_url,
            business_address=business.address or "",
            business_phone=business.phone or "",
            business_website=business.website or "",
            tracking_pixel_url=pixel_url,
            click_tracking_url=click_url,
            business_reply_email=business.email,
        )
        campaign.status = "sent" if success else "failed"
    elif customer.phone:
        success = send_sms(customer.phone, message)
        campaign.status = "sent" if success else "failed"
    return campaign

def run_automations(business_id):
    """Run all active automation rules for a business."""
    now = datetime.utcnow()
    last = _last_automation_run.get(business_id)
    if last and (now - last).total_seconds() < 3600:  # throttle to once per hour per business
        return
    _last_automation_run[business_id] = now
    try:
        today = now
        rules = AutomationRule.query.filter_by(business_id=business_id, active=True).all()
        b = Business.query.get(business_id)
        if not b:
            return
        for rule in rules:
            # Only run each rule once per day
            if rule.last_run and (today - rule.last_run).total_seconds() < 82800:
                continue

            if rule.rule_type == "birthday":
                customers = Customer.query.filter_by(business_id=business_id).filter(
                    Customer.dob != None, Customer.dob != "", Customer.unsubscribed == False
                ).all()
                for c in customers:
                    try:
                        dob = datetime.strptime(c.dob, "%Y-%m-%d")
                        bday = dob.replace(year=today.year)
                        if bday < today.replace(hour=0, minute=0, second=0):
                            bday = bday.replace(year=today.year + 1)
                        days_until = (bday - today).days
                        if days_until != rule.days_threshold:
                            continue
                        # Check not already sent this year
                        already_sent = Campaign.query.filter_by(
                            business_id=business_id,
                            customer_email=c.email or "",
                            campaign_type="birthday",
                        ).filter(Campaign.created_at >= today.replace(month=1, day=1, hour=0, minute=0, second=0)).first()
                        if already_sent:
                            continue
                        msg = rule.message_template or generate_ai_message(c.first_name, b.business_name, "birthday")
                        msg = msg.replace("{name}", c.first_name).replace("{business}", b.business_name)
                        _send_auto_campaign(b, c, "birthday", msg)
                    except Exception:
                        pass

            elif rule.rule_type == "reengagement":
                cutoff = datetime(today.year, today.month, today.day) - timedelta(days=rule.days_threshold)
                customers = Customer.query.filter_by(business_id=business_id, unsubscribed=False).all()
                for c in customers:
                    last_campaign = Campaign.query.filter_by(
                        business_id=business_id,
                        customer_email=c.email or "",
                    ).filter(Campaign.status == "sent").order_by(Campaign.created_at.desc()).first()
                    if last_campaign and last_campaign.created_at > cutoff:
                        continue
                    if not last_campaign and c.created_at > cutoff:
                        continue
                    msg = rule.message_template or generate_ai_message(c.first_name, b.business_name, "come_back")
                    msg = msg.replace("{name}", c.first_name).replace("{business}", b.business_name)
                    _send_auto_campaign(b, c, "come_back", msg)

            elif rule.rule_type == "welcome":
                # Send welcome to customers added in the last 24h who haven't received one
                cutoff_welcome = datetime(today.year, today.month, today.day, today.hour, today.minute) - timedelta(hours=24)
                new_customers = Customer.query.filter_by(business_id=business_id, unsubscribed=False).filter(
                    Customer.created_at >= cutoff_welcome
                ).all()
                for c in new_customers:
                    already = Campaign.query.filter_by(
                        business_id=business_id,
                        customer_email=c.email or "",
                        campaign_type="loyalty",
                    ).filter(Campaign.created_at >= c.created_at).first()
                    if already:
                        continue
                    msg = rule.message_template or f"Welcome to {b.business_name}, {c.first_name}! 🎉 We're so glad to have you. Look forward to exclusive offers just for you."
                    msg = msg.replace("{name}", c.first_name).replace("{business}", b.business_name)
                    _send_auto_campaign(b, c, "loyalty", msg)

            rule.last_run = today
        db.session.commit()
    except Exception as e:
        print(f"Automation error: {e}")


def enroll_customer_in_journeys(customer, trigger="signup"):
    """Enroll a customer in all active journeys with matching trigger."""
    try:
        now = datetime.utcnow()
        journeys = Journey.query.filter_by(business_id=customer.business_id, trigger=trigger, active=True).all()
        for journey in journeys:
            existing = CustomerJourney.query.filter_by(journey_id=journey.id, customer_id=customer.id).first()
            if existing:
                continue
            first_step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=0).first()
            delay = first_step.delay_days if first_step else 0
            cj = CustomerJourney(
                journey_id=journey.id,
                customer_id=customer.id,
                business_id=customer.business_id,
                next_step_order=0,
                enrolled_at=now,
                next_step_at=now + timedelta(days=delay)
            )
            db.session.add(cj)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Journey enroll error: {e}")


def process_journeys(business_id=None):
    """Run all due journey steps. Called by cron job daily."""
    now = datetime.utcnow()
    try:
        query = CustomerJourney.query.filter_by(completed=False, active=True).filter(CustomerJourney.next_step_at <= now)
        if business_id:
            query = query.filter_by(business_id=business_id)
        due = query.all()
        for cj in due:
            journey = Journey.query.get(cj.journey_id)
            if not journey or not journey.active:
                cj.active = False
                continue
            customer = Customer.query.get(cj.customer_id)
            if not customer or customer.unsubscribed:
                cj.active = False
                continue
            business = Business.query.get(cj.business_id)
            if not business:
                continue
            profile = BusinessProfile.query.filter_by(business_id=cj.business_id).first()

            # --- Goal achievement: exit journey early if goal met ---
            goal = getattr(journey, 'goal', 'none') or 'none'
            if goal == 'visit' and customer.last_visit and (now - customer.last_visit).days <= 3:
                cj.completed = True
                continue
            if goal == 'email_open' and cj.last_campaign_id:
                gc = Campaign.query.get(cj.last_campaign_id)
                if gc and gc.open_count and gc.open_count > 0:
                    cj.completed = True
                    continue

            # --- Frequency cap: skip if customer received too many messages this week ---
            freq_cap = getattr(journey, 'freq_cap_per_week', 0) or 0
            if freq_cap > 0:
                week_ago = now - timedelta(days=7)
                recent = Campaign.query.filter(
                    Campaign.business_id == cj.business_id,
                    Campaign.customer_email == customer.email,
                    Campaign.created_at >= week_ago,
                    Campaign.status == 'sent'
                ).count()
                if recent >= freq_cap:
                    cj.next_step_at = now + timedelta(days=1)
                    continue

            # --- Open-check phase: was waiting to see if email was opened ---
            if cj.waiting_for_open and cj.open_check_at and cj.open_check_at <= now:
                cj.waiting_for_open = False
                if cj.last_campaign_id:
                    sent_campaign = Campaign.query.get(cj.last_campaign_id)
                    if sent_campaign and sent_campaign.open_count == 0 and customer.phone:
                        # Email was not opened — send SMS fallback using same message
                        sms_msg = sent_campaign.message
                        send_sms(customer.phone, sms_msg)
                # Advance to next step
                next_step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=cj.next_step_order).first()
                if next_step:
                    cj.next_step_at = now + timedelta(days=next_step.delay_days)
                else:
                    cj.completed = True
                continue

            step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=cj.next_step_order).first()
            if not step:
                cj.completed = True
                continue
            # Check condition
            skip = False
            if step.condition == "visited":
                ref = customer.last_visit or customer.created_at
                if (now - ref).days > 7:
                    skip = True
            elif step.condition == "not_visited":
                ref = customer.last_visit or customer.created_at
                if (now - ref).days <= 7:
                    skip = True
            sent_campaign = None
            if not skip:
                lang = profile.language if profile else "English"
                cuisine = profile.cuisine_type if profile else ""
                dish = profile.signature_dish if profile else ""
                offer = profile.special_offer if profile else ""
                tone = profile.tone if profile else "friendly"
                if step.use_ai:
                    msg = generate_ai_message(customer.first_name, business.business_name, step.message_type,
                                              language=lang, cuisine=cuisine, dish=dish, offer=offer, tone=tone)
                else:
                    msg = (step.message_text or "").replace("{name}", customer.first_name).replace("{business}", business.business_name)
                if msg:
                    sent_campaign = _send_auto_campaign(business, customer, step.message_type, msg)
            # Advance — if email step with SMS fallback, enter open-check mode
            if (not skip and sent_campaign and step.channel == "email"
                    and getattr(step, "sms_fallback", False) and customer.phone):
                cj.last_campaign_id = sent_campaign.id
                cj.waiting_for_open = True
                cj.open_check_at = now + timedelta(days=step.sms_fallback_days or 2)
                cj.next_step_order += 1
                # next_step_at is controlled by open_check_at; set it far enough to avoid premature re-fire
                cj.next_step_at = cj.open_check_at
            else:
                cj.next_step_order += 1
                next_step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=cj.next_step_order).first()
                if next_step:
                    cj.next_step_at = now + timedelta(days=next_step.delay_days)
                else:
                    cj.completed = True

        # --- Re-entry: re-enroll completed customers if allow_reentry is on ---
        reentry_journeys = Journey.query.filter_by(active=True).filter(
            Journey.allow_reentry == True
        ).all()
        if business_id:
            reentry_journeys = [j for j in reentry_journeys if j.business_id == business_id]
        for journey in reentry_journeys:
            completed_cjs = CustomerJourney.query.filter_by(
                journey_id=journey.id, completed=True, active=True
            ).all()
            for cj in completed_cjs:
                # Re-enroll only if 30+ days since last completion
                if cj.next_step_at and (now - cj.next_step_at).days < 30:
                    continue
                first_step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=0).first()
                cj.next_step_order = 0
                cj.completed = False
                cj.waiting_for_open = False
                cj.last_campaign_id = None
                cj.enrolled_at = now
                cj.next_step_at = now + timedelta(days=first_step.delay_days if first_step else 0)

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Journey process error: {e}")


def _run_profile_autopilot(b, profile):
    """Run AI autopilot based on BusinessProfile settings. Called on dashboard load (throttled)."""
    if not profile or not profile.setup_complete:
        return
    now = datetime.utcnow()
    today_str = now.strftime("%Y-%m-%d")
    dish     = profile.signature_dish or "our specials"
    offer    = profile.special_offer  or "a special offer"
    tone     = profile.tone           or "friendly"
    language = profile.language       or "English"

    def ai_msg(name, campaign_hint, extra_ctx=""):
        try:
            return generate_ai_message(
                name, b.business_name, campaign_hint,
                cuisine=profile.cuisine_type or "",
                dish=dish, offer=offer, tone=tone,
                language=language, extra=extra_ctx
            )
        except Exception:
            return None

    def already_sent(customer, camp_type, since_days=30):
        cutoff = now - timedelta(days=since_days)
        return Campaign.query.filter_by(
            business_id=b.id,
            customer_email=customer.email or f"phone:{customer.phone}",
            campaign_type=camp_type,
        ).filter(Campaign.created_at >= cutoff).first() is not None

    customers = Customer.query.filter_by(business_id=b.id, unsubscribed=False).all()

    for c in customers:
        # ── REVIEW REQUEST: send review_days after signup ──
        if profile.auto_review and profile.google_review_url and c.email:
            days_since_join = (now - c.created_at).days
            if days_since_join == (profile.review_days or 3):
                if not already_sent(c, "review_request", since_days=365):
                    review_link = profile.google_review_url
                    msg = (
                        f"Hi {c.first_name}! 🌟\n\n"
                        f"Thank you so much for visiting {b.business_name}. "
                        f"We hope you had a wonderful experience!\n\n"
                        f"Would you mind taking 30 seconds to leave us a review? "
                        f"It means the world to a small business like ours.\n\n"
                        f"👉 {review_link}\n\n"
                        f"Thank you so much, {c.first_name}! 🙏"
                    )
                    _send_auto_campaign(b, c, "review_request", msg)

        # ── WIN-BACK: 30 days since last visit or signup ──
        if profile.auto_winback and (c.email or c.phone):
            ref_date = c.last_visit or c.created_at
            days_inactive = (now - ref_date).days
            if days_inactive >= 30:
                if not already_sent(c, "come_back", since_days=30):
                    msg = ai_msg(c.first_name, "come_back", f"customer hasn't visited in {days_inactive} days")
                    if msg:
                        _send_auto_campaign(b, c, "come_back", msg)

        # ── LOYALTY MILESTONE: every N visits ──
        if profile.auto_loyalty and c.email:
            n = profile.loyalty_reward_visits or 5
            if c.visit_count > 0 and c.visit_count % n == 0:
                if not already_sent(c, "loyalty", since_days=7):
                    msg = ai_msg(
                        c.first_name, "loyalty",
                        f"customer just reached {c.visit_count} visits — send a loyalty reward"
                    )
                    if msg:
                        _send_auto_campaign(b, c, "loyalty", msg)

    # ── BIRTHDAY: send on the customer's birthday ──
    if profile.auto_birthday:
        for c in customers:
            if not c.dob or not (c.email or c.phone):
                continue
            try:
                dob = datetime.strptime(c.dob[:10], "%Y-%m-%d")
                if dob.month == now.month and dob.day == now.day:
                    if not already_sent(c, "birthday", since_days=300):
                        msg = ai_msg(c.first_name, "birthday", f"it is {c.first_name}'s birthday today")
                        if msg:
                            _send_auto_campaign(b, c, "birthday", msg)
            except Exception:
                pass

    # ── WEEKLY SPECIAL: on the configured send day ──
    if profile.auto_weekly:
        day_name = now.strftime("%A")  # Monday, Tuesday, …
        send_day = profile.weekly_send_day or "Tuesday"
        if day_name == send_day:
            # Check if we already sent weekly special today
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            already_today = Campaign.query.filter_by(
                business_id=b.id, campaign_type="weekly_special"
            ).filter(Campaign.created_at >= today_start).first()
            if not already_today:
                active_customers = [c for c in customers if c.email or c.phone]
                for c in active_customers[:200]:  # cap at 200 per run
                    msg = ai_msg(c.first_name, "weekend", f"send weekly {send_day} special")
                    if msg:
                        _send_auto_campaign(b, c, "weekly_special", msg)

    # ── FLASH DEAL: on slow days ──
    if profile.auto_flash and profile.slow_days:
        day_name = now.strftime("%A")
        slow = [d.strip() for d in profile.slow_days.split(",")]
        if day_name in slow:
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            already_today = Campaign.query.filter_by(
                business_id=b.id, campaign_type="flash_deal"
            ).filter(Campaign.created_at >= today_start).first()
            if not already_today:
                active_customers = [c for c in customers if c.email or c.phone]
                for c in active_customers[:200]:
                    msg = ai_msg(c.first_name, "promotion", f"urgent flash deal — slow {day_name}, fill seats now")
                    if msg:
                        _send_auto_campaign(b, c, "flash_deal", msg)


def _maybe_send_weekly_summary(b, profile, today):
    """Send a weekly summary email to the business owner every Monday, once per week."""
    if today.weekday() != 0:   # 0 = Monday
        return
    if profile.last_weekly_summary:
        days_since = (today - profile.last_weekly_summary).days
        if days_since < 6:
            return
    week_ago = today - timedelta(days=7)
    sent_week   = Campaign.query.filter_by(business_id=b.id, status="sent").filter(Campaign.created_at >= week_ago).count()
    opens_week  = sum(c.open_count or 0 for c in Campaign.query.filter_by(business_id=b.id, status="sent").filter(Campaign.created_at >= week_ago).all())
    new_customers = Customer.query.filter_by(business_id=b.id).filter(Customer.created_at >= week_ago).count()
    total_customers = Customer.query.filter_by(business_id=b.id).count()

    active_autos = []
    if profile.auto_welcome:  active_autos.append("Welcome Message")
    if profile.auto_weekly:   active_autos.append("Weekly Special")
    if profile.auto_flash:    active_autos.append("Flash Deal")
    if profile.auto_birthday: active_autos.append("Birthday Offer")
    if profile.auto_winback:  active_autos.append("Win-Back")
    auto_list = ", ".join(active_autos) if active_autos else "None active"

    subject = f"📊 Your weekly Revvio report — {b.business_name}"
    body = (
        f"Hi {b.owner_name},\n\n"
        f"Here's what your AI did for {b.business_name} this week:\n\n"
        f"  📨 Messages sent:     {sent_week}\n"
        f"  👁 Emails opened:     {opens_week}\n"
        f"  🆕 New customers:     {new_customers}\n"
        f"  👥 Total customers:   {total_customers}\n\n"
        f"Active automations: {auto_list}\n\n"
        f"Your AI is working 24/7 so you don't have to. 🚀\n\n"
        f"— The Revvio Team\n"
        f"View your dashboard: https://revvio.ai/dashboard"
    )
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;padding:24px;background:#f9f9f9;">
      <div style="background:linear-gradient(135deg,#667eea,#764ba2);border-radius:14px;padding:28px;text-align:center;margin-bottom:20px;">
        <h1 style="color:#fff;margin:0;font-size:22px;">Weekly Report 📊</h1>
        <p style="color:rgba(255,255,255,0.85);margin:6px 0 0;">{b.business_name}</p>
      </div>
      <div style="background:#fff;border-radius:14px;padding:24px;margin-bottom:16px;">
        <p style="color:#555;margin-top:0;">Hi <strong>{b.owner_name}</strong>, here's what your AI did this week:</p>
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:10px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:14px;">📨 Messages sent</td><td style="text-align:right;font-size:20px;font-weight:800;color:#667eea;">{sent_week}</td></tr>
          <tr><td style="padding:10px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:14px;">👁 Emails opened</td><td style="text-align:right;font-size:20px;font-weight:800;color:#22c55e;">{opens_week}</td></tr>
          <tr><td style="padding:10px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:14px;">🆕 New customers</td><td style="text-align:right;font-size:20px;font-weight:800;color:#f59e0b;">{new_customers}</td></tr>
          <tr><td style="padding:10px 0;color:#888;font-size:14px;">👥 Total customers</td><td style="text-align:right;font-size:20px;font-weight:800;color:#333;">{total_customers}</td></tr>
        </table>
      </div>
      <div style="background:#fff;border-radius:14px;padding:18px 24px;margin-bottom:16px;">
        <p style="color:#888;font-size:13px;margin:0;">🤖 Active automations: <strong style="color:#667eea;">{auto_list}</strong></p>
      </div>
      <div style="text-align:center;margin-top:20px;">
        <a href="https://revvio.ai/dashboard" style="background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;padding:13px 28px;border-radius:10px;display:inline-block;font-weight:700;text-decoration:none;font-size:15px;">View Dashboard →</a>
      </div>
      <p style="text-align:center;color:#bbb;font-size:11px;margin-top:20px;">Revvio · Your AI is working 24/7 so you don't have to.</p>
    </div>
    """
    try:
        send_email(b.email, subject, body, customer_name=b.owner_name,
                   business_name="Revvio", campaign_type="promotion", html_override=html)
        profile.last_weekly_summary = today
        db.session.commit()
    except Exception as e:
        print(f"Weekly summary error: {e}")


def build_html_email(business_name, customer_name, message, campaign_type, unsubscribe_url="", business_address="", business_phone="", business_website="", tracking_pixel_url="", click_tracking_url=""):
    unsub_html = ""
    if unsubscribe_url:
        unsub_html = f' · <a href="{unsubscribe_url}" style="color:#aaa;font-size:11px;">Unsubscribe</a>'
    contact_parts = []
    if business_address:
        contact_parts.append(business_address)
    if business_phone:
        contact_parts.append(business_phone)
    if business_website:
        site = business_website if business_website.startswith("http") else f"https://{business_website}"
        contact_parts.append(f'<a href="{site}" style="color:#aaa;">{business_website}</a>')
    contact_line = " · ".join(contact_parts)
    cta_labels = {
        "come_back":  "Come Back Today!",
        "weekend":    "Visit Us This Weekend!",
        "lunch":      "Join Us for Lunch!",
        "dinner":     "Reserve Your Table!",
        "birthday":   "Claim Your Birthday Gift!",
        "loyalty":    "Claim Your Reward!",
        "happy_hour": "Join Happy Hour!",
        "new_item":   "Try It Today!",
        "promotion":  "Claim This Offer!",
    }
    cta_text = cta_labels.get(campaign_type, "Visit Us Today!")
    header_labels = {
        "come_back":  "We Miss You!",
        "weekend":    "Weekend Special",
        "lunch":      "Lunch Special",
        "dinner":     "Dinner Special",
        "birthday":   "Happy Birthday!",
        "loyalty":    "You're a VIP!",
        "happy_hour": "Happy Hour!",
        "new_item":   "Something New!",
        "promotion":  "Special Offer",
    }
    header_text = header_labels.get(campaign_type, "Special Offer Just For You")
    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;background:#f5f5f5;">
    <div style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:30px;border-radius:12px;text-align:center;margin-bottom:20px;">
        <h1 style="color:white;margin:0;font-size:26px;">{business_name}</h1>
        <p style="color:rgba(255,255,255,0.9);margin:8px 0 0 0;">{header_text}</p>
    </div>
    <div style="background:white;padding:30px;border-radius:12px;margin-bottom:20px;">
        <p style="font-size:16px;color:#333;">Hi <strong>{customer_name}</strong>,</p>
        <p style="font-size:16px;color:#555;line-height:1.7;">{message}</p>
        <div style="text-align:center;margin-top:25px;">
            {'<a href="' + click_tracking_url + '" style="background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:14px 30px;border-radius:8px;display:inline-block;font-size:16px;font-weight:bold;text-decoration:none;">' + cta_text + '</a>' if click_tracking_url else '<p style="background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:14px 30px;border-radius:8px;display:inline-block;font-size:16px;font-weight:bold;">' + cta_text + '</p>'}
        </div>
        {f'<p style="text-align:center;color:#888;font-size:13px;margin-top:20px;">📍 {business_address}</p>' if business_address else ''}
        {f'<p style="text-align:center;color:#888;font-size:13px;margin:4px 0;">📞 {business_phone}</p>' if business_phone else ''}
    </div>
    <p style="text-align:center;color:#999;font-size:12px;">
        You received this because you're a valued customer of {business_name}.<br>
        {business_name} · {contact_line}{unsub_html}
    </p>
    {f'<img src="{tracking_pixel_url}" width="1" height="1" style="display:none;" alt="">' if tracking_pixel_url else ''}
    </body></html>
    """

def send_email(to_email, subject, body, customer_name="", business_name="", campaign_type="promotion", unsubscribe_url="", business_address="", business_phone="", business_website="", tracking_pixel_url="", click_tracking_url="", business_reply_email="", html_override=None):
    from_name = business_name or "Revvio"
    html_body = html_override if html_override else build_html_email(from_name, customer_name or "Valued Customer", body, campaign_type, unsubscribe_url, business_address, business_phone, business_website, tracking_pixel_url, click_tracking_url)

    sendgrid_key = os.getenv("SENDGRID_API_KEY")
    if sendgrid_key:
        # Option 2: SendGrid — sends from mail.revvio.ai with business name
        try:
            import sendgrid as sg_module
            from sendgrid.helpers.mail import Mail, Email, To, ReplyTo
            from_email = os.getenv("SENDGRID_FROM_EMAIL", "noreply@mail.revvio.ai")
            message = Mail(
                from_email=Email(from_email, from_name),
                to_emails=To(to_email),
                subject=subject,
                plain_text_content=body,
                html_content=html_body,
            )
            if business_reply_email:
                message.reply_to = ReplyTo(business_reply_email, from_name)
            if unsubscribe_url:
                message.header = {"List-Unsubscribe": f"<{unsubscribe_url}>"}
            sg = sg_module.SendGridAPIClient(api_key=sendgrid_key)
            response = sg.send(message)
            return response.status_code in (200, 202)
        except Exception as e:
            print(f"SendGrid error: {e}")
            return False
    else:
        # Option 1: SMTP — From name is business name, Reply-To is business email
        try:
            smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
            smtp_port = int(os.getenv("SMTP_PORT", 587))
            sender_email = os.getenv("SENDER_EMAIL")
            sender_password = os.getenv("SENDER_PASSWORD")
            if not all([sender_email, sender_password]):
                print("Email not configured")
                return False
            msg = MIMEMultipart("alternative")
            msg["From"] = f'"{from_name}" <{sender_email}>'
            msg["To"] = to_email
            msg["Subject"] = subject
            if business_reply_email:
                msg["Reply-To"] = business_reply_email
            if unsubscribe_url:
                msg["List-Unsubscribe"] = f"<{unsubscribe_url}>"
            msg.attach(MIMEText(body, "plain"))
            msg.attach(MIMEText(html_body, "html"))
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)
            return True
        except Exception as e:
            print(f"Email error: {e}")
            return False

def send_sms(to_phone, message):
    """Send SMS with TCPA compliance — mandatory STOP footer."""
    if not twilio_client or not twilio_phone:
        return False
    # Append mandatory opt-out footer if not already present
    body = message.rstrip() + (" Reply STOP to unsubscribe." if "STOP" not in message else "")
    try:
        twilio_client.messages.create(body=body, from_=twilio_phone, to=to_phone)
        return True
    except Exception as e:
        print(f"SMS error: {e}")
        return False

def generate_ai_message(customer_name, business_name, campaign_type, raise_on_error=False, language="English", cuisine="", dish="", offer="", tone="friendly", extra=""):
    fallbacks = {
        "come_back":  f"We miss you, {customer_name}! Come back to {business_name} and enjoy 15% off your next visit. We'd love to see you again!",
        "weekend":    f"Hi {customer_name}! Weekend special at {business_name} — amazing food and great deals this weekend only!",
        "lunch":      f"Hi {customer_name}! Join us for lunch at {business_name} today. Fresh food, great prices, and a warm welcome!",
        "dinner":     f"Hi {customer_name}! Special dinner offer tonight at {business_name}. Reserve your table and enjoy an unforgettable evening!",
        "birthday":   f"Happy Birthday {customer_name}! Enjoy 20% off your next visit to {business_name}. Use code: BIRTHDAY20",
        "loyalty":    f"Thank you for being a loyal customer, {customer_name}! Your exclusive reward is waiting at {business_name}.",
        "happy_hour":    f"Hi {customer_name}! Happy Hour at {business_name} — amazing drinks and bites at special prices. Come join us!",
        "new_item":      f"Hi {customer_name}! We just launched something exciting at {business_name}. Come be the first to try it!",
        "promotion":     f"Hi {customer_name}! Special promotion at {business_name} just for you. Don't miss out — visit us soon!",
        "hotel_stay":    f"Hi {customer_name}! Exclusive stay offer at {business_name}. Book now and save on your next visit!",
        "hotel_upgrade": f"Hi {customer_name}! Enjoy a complimentary room upgrade on your next stay at {business_name}. Limited availability!",
        "hotel_event":   f"Hi {customer_name}! Special event package at {business_name} — the perfect getaway. Book before it's gone!",
        "hotel_loyalty": f"Thank you for being a loyal guest, {customer_name}! You've earned exclusive rewards at {business_name}.",
    }

    if not client:
        return fallbacks.get(campaign_type, fallbacks["promotion"])

    prompts = {
        "come_back":     f"Write a warm 'we miss you, come back' offer for {customer_name} from {business_name}. Include a discount to return.",
        "weekend":       f"Write a weekend special promotion for {customer_name} from {business_name}. Make it exciting.",
        "lunch":         f"Write a lunch deal promotion for {customer_name} from {business_name}. Make it appetizing.",
        "dinner":        f"Write a dinner special for {customer_name} from {business_name}. Make it feel exclusive and special.",
        "birthday":      f"Write a birthday offer for {customer_name} from {business_name}. Include a discount.",
        "loyalty":       f"Write a loyalty reward message for {customer_name} from {business_name}. Thank them warmly.",
        "happy_hour":    f"Write a happy hour promotion for {customer_name} from {business_name}. Make it fun.",
        "new_item":      f"Write a new menu item announcement for {customer_name} from {business_name}. Make it exciting.",
        "promotion":     f"Write a general promotion for {customer_name} from {business_name}. Make it compelling.",
        "hotel_stay":    f"Write a special hotel stay offer for {customer_name} from {business_name} hotel. Include a discount on room rate.",
        "hotel_upgrade": f"Write a room upgrade offer for {customer_name} from {business_name} hotel. Make it feel luxurious and exclusive.",
        "hotel_event":   f"Write an event/package deal promotion for {customer_name} from {business_name} hotel. Make it sound exciting.",
        "hotel_loyalty": f"Write a guest loyalty reward message for {customer_name} from {business_name} hotel. Thank them warmly.",
    }

    context_parts = []
    if cuisine:  context_parts.append(f"cuisine: {cuisine}")
    if dish:     context_parts.append(f"signature dish: {dish}")
    if offer:    context_parts.append(f"current offer: {offer}")
    if extra:    context_parts.append(extra)
    context_str = (". Context: " + ", ".join(context_parts)) if context_parts else ""
    lang_str = f" Write the message in {language}." if language and language != "English" else ""
    tone_str = f" Use a {tone} tone." if tone else ""

    prompt = prompts.get(campaign_type, prompts["promotion"]) + context_str + tone_str + lang_str + " Keep it under 3 sentences. No markdown. Include emojis."

    try:
        return clean_ai_text(_call_openai(prompt))
    except Exception as e:
        print(f"AI error: {e}")
        if raise_on_error:
            raise
        return fallbacks.get(campaign_type, fallbacks["promotion"])

# ============================================
# ROUTES
# ============================================

@app.route("/try", methods=["GET", "POST"])
def try_demo():
    """Interactive AI demo — enter restaurant details, get real AI messages."""
    messages = None
    form_data = {}
    if request.method == "POST":
        business_name = request.form.get("business_name", "").strip()
        cuisine       = request.form.get("cuisine", "").strip()
        dish          = request.form.get("dish", "").strip()
        offer         = request.form.get("offer", "").strip()
        customer_name = request.form.get("customer_name", "Sarah").strip() or "Sarah"
        form_data = {"business_name": business_name, "cuisine": cuisine, "dish": dish, "offer": offer, "customer_name": customer_name}
        if business_name:
            def gen(ctype, hint):
                try:
                    prompt = (
                        f"You are a marketing AI for {business_name}, a {cuisine} restaurant. "
                        f"Their signature dish is '{dish}'. Their current offer is '{offer}'. "
                        f"Write a {hint} for a customer named {customer_name}. "
                        f"Keep it under 3 sentences. No markdown. Use emojis. Sound warm and genuine."
                    )
                    return clean_ai_text(_call_openai(prompt, max_tokens=120))
                except Exception:
                    fallbacks2 = {
                        "welcome":  f"Welcome to {business_name}, {customer_name}! 🎉 We're so glad you're here. Your first visit earns you {offer or 'a special treat'} — can't wait to see you!",
                        "birthday": f"Happy Birthday {customer_name}! 🎂 The whole team at {business_name} wishes you an amazing day. Come celebrate with us and enjoy {offer or '20% off'}!",
                        "winback":  f"Hey {customer_name}, we miss you! 💔 It's been a while since you visited {business_name}. Come back this week and enjoy {offer or '15% off'} — just for you.",
                        "weekly":   f"Hi {customer_name}! 📅 This week's special at {business_name}: our famous {dish or 'chef special'} with {offer or 'an exclusive deal'}. Come in before it's gone!",
                    }
                    return fallbacks2.get(ctype, "")
            messages = {
                "welcome":  gen("welcome",  f"warm welcome message for a brand new customer"),
                "birthday": gen("birthday", f"birthday offer message"),
                "winback":  gen("winback",  f"'we miss you, come back' offer for a customer who hasn't visited in 30 days"),
                "weekly":   gen("weekly",   f"weekly Tuesday special promotion highlighting {dish or 'their best dish'} and {offer or 'a deal'}"),
            }
    return render_template("try_demo.html", messages=messages, form_data=form_data)


@app.route("/test")
def test():
    return f"<h1>Flask is working!</h1><p>Environment: {ENV}</p>"

@app.route("/test-ai")
def test_ai():
    key = os.getenv("OPENAI_API_KEY", "")
    if not key:
        return "<h2>❌ OPENAI_API_KEY is not set in environment variables.</h2>"
    if not key.startswith("sk-"):
        return f"<h2>❌ Key looks wrong — starts with '{key[:6]}...' (should start with 'sk-')</h2>"
    if not client:
        return f"<h2>❌ OpenAI client failed to initialize.</h2><pre>Error: {openai_init_error}</pre><p>Key starts with: {key[:12]}...</p><p>Key length: {len(key)} chars</p>"
    try:
        result = _call_openai("Say hello in one word.", max_tokens=10)
        return f"<h2>✅ OpenAI is working!</h2><p>Response: <strong>{result}</strong></p><p>Key starts with: {key[:8]}...</p>"
    except Exception as e:
        return f"<h2>❌ OpenAI call failed</h2><pre>{str(e)}</pre><p>Key starts with: {key[:8]}...</p>"

@app.route("/", methods=["GET"])
def landing():
    try:
        return render_template("landing.html", google_client_id=os.getenv("GOOGLE_CLIENT_ID", ""))
    except Exception as e:
        return f"<h1>Error: {e}</h1>", 500

@app.route("/privacy")
def privacy():
    return render_template("privacy.html")

@app.route("/terms")
def terms():
    return render_template("terms.html")

@app.route("/contact", methods=["GET", "POST"])
def contact():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        message = request.form.get("message", "").strip()
        if not all([name, email, message]):
            flash("All fields required.", "error")
            return redirect("/contact")
        try:
            db.session.add(ContactMessage(name=name, email=email, message=message))
            db.session.commit()
            flash("Message sent! We'll get back to you within a few hours.", "success")
            return redirect("/contact")
        except:
            db.session.rollback()
            flash("Error saving message.", "error")
    return render_template("contact.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if "user_id" in session and not session.get("is_demo"):
        return redirect("/dashboard")
    if request.method == "POST":
        business_name = request.form.get("business_name", "").strip()
        owner_name = request.form.get("owner_name", "").strip()
        email = request.form.get("email", "").strip()
        raw_password = request.form.get("password", "")
        if not all([business_name, owner_name, email, raw_password]):
            flash("All fields required.", "error")
            return redirect("/register")
        if len(raw_password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return redirect("/register")
        if Business.query.filter_by(email=email).first():
            flash("Email already registered.", "error")
            return redirect("/login")
        session.pop("is_demo", None)
        session.pop("user_id", None)
        hashed = bcrypt.hashpw(raw_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        b = Business(business_name=business_name, owner_name=owner_name, email=email, password=hashed)
        try:
            db.session.add(b)
            db.session.commit()
            session["user_id"] = b.id
            session.permanent = True
            # Send welcome email to new user
            try:
                welcome_html = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;padding:24px;">
      <div style="background:linear-gradient(135deg,#667eea,#764ba2);border-radius:14px;padding:28px;text-align:center;margin-bottom:20px;">
        <h1 style="color:#fff;margin:0;font-size:24px;">Welcome to Revvio! 🚀</h1>
        <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;">AI Marketing Auto-Pilot for {business_name}</p>
      </div>
      <div style="background:#f9f9f9;border-radius:14px;padding:24px;margin-bottom:16px;">
        <p style="color:#333;font-size:16px;">Hi <strong>{owner_name}</strong>! 👋</p>
        <p style="color:#555;line-height:1.7;">You're just 2 minutes away from having AI handle all your restaurant marketing automatically.</p>
        <p style="color:#555;line-height:1.7;"><strong>What happens next:</strong></p>
        <ul style="color:#555;line-height:2;">
          <li>Complete your AI setup (takes 2 min)</li>
          <li>Share your customer signup link</li>
          <li>AI starts sending welcome messages, birthday offers, and weekly specials — automatically</li>
        </ul>
      </div>
      <div style="text-align:center;margin-top:20px;">
        <a href="https://revvio.ai/onboarding" style="background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;padding:14px 32px;border-radius:10px;display:inline-block;font-weight:700;text-decoration:none;font-size:16px;">Complete My AI Setup →</a>
      </div>
      <p style="text-align:center;color:#bbb;font-size:12px;margin-top:20px;">Revvio · hello@revvio.ai · Plano, TX</p>
    </div>
    """
                send_email(
                    email,
                    f"Welcome to Revvio, {owner_name}! 🚀 Let's activate your AI",
                    f"Welcome to Revvio, {owner_name}! Complete your AI setup at https://revvio.ai/onboarding",
                    customer_name=owner_name,
                    business_name="Revvio",
                    campaign_type="promotion",
                    html_override=welcome_html
                )
            except Exception:
                pass
            flash(f"Welcome {owner_name}! Let's get your AI set up in 2 minutes. 🚀", "success")
            return redirect("/onboarding")
        except:
            db.session.rollback()
            flash("Error creating account.", "error")
    return render_template("index.html", google_client_id=os.getenv("GOOGLE_CLIENT_ID", ""))

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session and not session.get("is_demo"):
        return redirect("/dashboard")
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        b = Business.query.filter_by(email=email).first()
        if b and bcrypt.checkpw(password.encode("utf-8"), b.password.encode("utf-8")):
            session["user_id"] = b.id
            session.permanent = True
            flash(f"Welcome back, {b.owner_name}!", "success")
            return redirect("/dashboard")
        flash("Invalid email or password.", "error")
        return redirect("/login")
    return render_template("login.html", google_client_id=os.getenv("GOOGLE_CLIENT_ID", ""))

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        b = Business.query.filter_by(email=email).first()
        if b:
            import secrets as _secrets
            new_pw = _secrets.token_urlsafe(10)
            b.password = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
            db.session.commit()
            try:
                send_email(
                    email,
                    "Your Revvio temporary password",
                    f"Hi {b.owner_name},\n\nYour temporary password is: {new_pw}\n\nLog in at https://revvio.ai/login and change it in Settings.\n\n— Revvio Team",
                    customer_name=b.owner_name,
                    business_name="Revvio",
                    campaign_type="promotion"
                )
            except Exception:
                pass
        # Always show success (don't reveal if email exists)
        flash("If that email is registered, you'll receive a temporary password shortly.", "success")
        return redirect("/login")
    return render_template("forgot_password.html")

@app.route("/auth/google", methods=["POST"])
def auth_google():
    """Verify Google ID token and log in or create account."""
    from google.oauth2 import id_token
    from google.auth.transport import requests as google_requests
    credential = request.form.get("credential") or (request.json or {}).get("credential", "")
    if not credential:
        flash("Google sign-in failed. Please try again.", "error")
        return redirect("/login")
    google_client_id = os.getenv("GOOGLE_CLIENT_ID", "")
    if not google_client_id:
        flash("Google Sign-In is not configured yet.", "error")
        return redirect("/login")
    try:
        info = id_token.verify_oauth2_token(credential, google_requests.Request(), google_client_id)
    except Exception as e:
        print(f"Google token error: {e}")
        flash("Google sign-in failed. Please try again.", "error")
        return redirect("/login")
    email      = info.get("email", "").strip().lower()
    name       = info.get("name", "").strip()
    first_name = info.get("given_name", name.split()[0] if name else "User").strip()
    if not email:
        flash("Could not get email from Google account.", "error")
        return redirect("/login")
    # Find or create business account
    b = Business.query.filter_by(email=email).first()
    if b:
        # Existing user — log in
        session["user_id"] = b.id
        session.permanent = True
        session.pop("is_demo", None)
        flash(f"Welcome back, {b.owner_name}! 👋", "success")
        return redirect("/dashboard")
    else:
        # New user — create account with a random password (Google handles auth)
        import secrets
        random_pw = secrets.token_hex(24)
        hashed = bcrypt.hashpw(random_pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        b = Business(
            business_name=f"{first_name}'s Restaurant",
            owner_name=name or first_name,
            email=email,
            password=hashed,
        )
        try:
            db.session.add(b)
            db.session.commit()
            session["user_id"] = b.id
            session.permanent = True
            session.pop("is_demo", None)
            try:
                send_email(
                    email,
                    f"Welcome to Revvio, {first_name}! 🚀 Let's activate your AI",
                    f"Welcome to Revvio! Complete your AI setup at https://revvio.ai/onboarding",
                    customer_name=first_name,
                    business_name="Revvio",
                    campaign_type="promotion"
                )
            except Exception:
                pass
            flash(f"Welcome to Revvio, {first_name}! Let's set up your AI in 2 minutes. 🚀", "success")
            return redirect("/onboarding")
        except Exception:
            db.session.rollback()
            flash("Error creating account. Please try again.", "error")
            return redirect("/register")


@app.route("/cron/run-automations")
def cron_run_automations():
    """
    Called daily by Render Cron Job (or any scheduler).
    Runs AI autopilot for every business that has completed setup.
    Protect with a secret token so only the scheduler can call it.
    """
    token = request.args.get("token", "")
    cron_secret = os.getenv("CRON_SECRET", "")
    if cron_secret and token != cron_secret:
        return jsonify({"error": "Unauthorized"}), 401

    today = datetime.utcnow()
    businesses = Business.query.all()
    results = []
    for b in businesses:
        if b.email == DEMO_EMAIL:
            continue  # skip demo account
        profile = BusinessProfile.query.filter_by(business_id=b.id).first()
        if not profile or not profile.setup_complete:
            continue
        try:
            _run_profile_autopilot(b, profile)
            _maybe_send_weekly_summary(b, profile, today)
            run_automations(b.id)
            process_journeys(b.id)
            db.session.commit()
            results.append({"business": b.business_name, "status": "ok"})
        except Exception as e:
            db.session.rollback()
            results.append({"business": b.business_name, "status": f"error: {e}"})

    return jsonify({"ran": len(results), "results": results, "timestamp": today.isoformat()})


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("is_demo", None)
    flash("You've been logged out.", "success")
    return redirect("/")

@app.route("/demo-logout")
def demo_logout():
    session.pop("user_id", None)
    session.pop("is_demo", None)
    if request.args.get("next") == "login":
        return redirect("/login")
    return redirect("/register")

@app.route("/dashboard")
def dashboard():
    b = current_business()
    if not b:
        return redirect("/login")
    process_scheduled_campaigns()
    run_automations(b.id)
    today = datetime.utcnow()

    # Redirect new businesses that haven't completed onboarding
    profile = BusinessProfile.query.filter_by(business_id=b.id).first()
    if not profile or not profile.setup_complete:
        return redirect("/onboarding")

    # Run AI autopilot (profile-based automations)
    _run_profile_autopilot(b, profile)

    # Send weekly summary every Monday (throttled — once per week)
    _maybe_send_weekly_summary(b, profile, today)

    total_customers = Customer.query.filter_by(business_id=b.id).count()
    total_campaigns = Campaign.query.filter_by(business_id=b.id).count()
    sent_campaigns = Campaign.query.filter_by(business_id=b.id, status="sent").count()
    campaigns = Campaign.query.filter_by(business_id=b.id).order_by(Campaign.created_at.desc()).all()
    customer_limit = get_plan_limit(b.plan, "customers")
    campaign_limit = get_plan_limit(b.plan, "campaigns_month")
    month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    campaigns_this_month = Campaign.query.filter_by(business_id=b.id).filter(Campaign.created_at >= month_start).count()

    # New customers this week
    week_ago = today - timedelta(days=7)
    new_this_week = Customer.query.filter_by(business_id=b.id).filter(Customer.created_at >= week_ago).count()

    # Open rate
    total_opens = sum((c.open_count or 0) for c in campaigns if c.status == "sent")
    open_rate = round(total_opens / sent_campaigns * 100) if sent_campaigns > 0 else 0

    # Campaigns by type for chart
    type_counts = {}
    for c in campaigns:
        if c.status == "sent":
            type_counts[c.campaign_type] = type_counts.get(c.campaign_type, 0) + 1
    ct = get_campaign_types()
    chart_labels = [ct.get(k, k) for k in type_counts]
    chart_data = list(type_counts.values())

    # Birthday radar — customers with birthdays in the next 7 days
    birthday_customers = []
    all_customers_dob = Customer.query.filter_by(business_id=b.id).filter(Customer.dob != None, Customer.dob != "").all()
    for c in all_customers_dob:
        try:
            dob = datetime.strptime(c.dob, "%Y-%m-%d")
            this_year_bday = dob.replace(year=today.year)
            if this_year_bday < today.replace(hour=0, minute=0, second=0):
                this_year_bday = this_year_bday.replace(year=today.year + 1)
            days_until = (this_year_bday - today).days
            if 0 <= days_until <= 7:
                birthday_customers.append({"customer": c, "days": days_until})
        except Exception:
            pass
    birthday_customers.sort(key=lambda x: x["days"])

    # Top engaged customers (most campaigns sent to them)
    from collections import Counter
    email_counter = Counter(c.customer_email for c in campaigns if c.status == "sent" and c.customer_email)
    top_emails = [e for e, _ in email_counter.most_common(5)]
    top_customers = []
    for email in top_emails:
        c = Customer.query.filter_by(business_id=b.id, email=email).first()
        if c:
            top_customers.append({"customer": c, "count": email_counter[email]})

    # Smart Insights — actionable suggestions based on data
    insights = []
    # 1. Inactive customers
    cutoff_60 = today - timedelta(days=60)
    all_active = Customer.query.filter_by(business_id=b.id, unsubscribed=False).all()
    inactive_count = 0
    for c in all_active:
        last = Campaign.query.filter_by(business_id=b.id, customer_email=c.email or "").filter(
            Campaign.status == "sent"
        ).order_by(Campaign.created_at.desc()).first()
        if not last or last.created_at < cutoff_60:
            inactive_count += 1
    if inactive_count > 0:
        insights.append({
            "icon": "😴", "color": "#f59e0b",
            "title": f"{inactive_count} guests haven't heard from you in 60+ days",
            "body": "Send a 'We miss you' offer to bring them back.",
            "action": "Re-engage Now", "url": "/bulk-send?segment=inactive"
        })
    # 2. Uncontacted customers (never received a campaign)
    never_contacted = sum(1 for c in all_active if not Campaign.query.filter_by(
        business_id=b.id, customer_email=c.email or ""
    ).first())
    if never_contacted > 0:
        insights.append({
            "icon": "👋", "color": "#667eea",
            "title": f"{never_contacted} guests have never received a message",
            "body": "Welcome them with an intro offer.",
            "action": "Send Welcome", "url": "/bulk-send?segment=new"
        })
    # 3. Best performing campaign type
    type_opens = {}
    type_sent_count = {}
    for c in campaigns:
        if c.status == "sent":
            type_sent_count[c.campaign_type] = type_sent_count.get(c.campaign_type, 0) + 1
            type_opens[c.campaign_type] = type_opens.get(c.campaign_type, 0) + (c.open_count or 0)
    best_type = None
    best_rate = 0
    for t, opens in type_opens.items():
        sent = type_sent_count.get(t, 0)
        rate = round(opens / sent * 100) if sent >= 3 else 0
        if rate > best_rate:
            best_rate = rate
            best_type = t
    if best_type and best_rate >= 20:
        insights.append({
            "icon": "🏆", "color": "#10b981",
            "title": f"Your '{ct.get(best_type, best_type)}' campaigns get {best_rate}% open rate",
            "body": "That's above average! Send more of what's working.",
            "action": "Create One Now", "url": f"/create-campaign"
        })
    # 4. Customers with no email (SMS only)
    sms_only = Customer.query.filter_by(business_id=b.id).filter(
        Customer.phone.isnot(None), Customer.phone != "",
        (Customer.email.is_(None)) | (Customer.email == "")
    ).count()
    if sms_only > 0:
        insights.append({
            "icon": "📱", "color": "#8b5cf6",
            "title": f"{sms_only} guests only have a phone number",
            "body": "Reach them via SMS — email won't work for these guests.",
            "action": "Send SMS", "url": "/quick-sms"
        })
    # 5. Customers missing birthday
    no_dob = Customer.query.filter_by(business_id=b.id).filter(
        (Customer.dob.is_(None)) | (Customer.dob == "")
    ).count()
    if no_dob > 0 and total_customers > 0:
        insights.append({
            "icon": "🎂", "color": "#ec4899",
            "title": f"{no_dob} guests are missing a birthday date",
            "body": "Add birthdays to unlock birthday automation and radar.",
            "action": "View Customers", "url": "/customers"
        })

    return render_template(
        "dashboard.html",
        business_name=b.business_name,
        total_customers=total_customers,
        total_campaigns=total_campaigns,
        sent_campaigns=sent_campaigns,
        open_rate=open_rate,
        new_this_week=new_this_week,
        plan=b.plan,
        campaigns=campaigns,
        customer_limit=customer_limit,
        campaign_limit=campaign_limit,
        campaigns_this_month=campaigns_this_month,
        campaign_types=ct,
        birthday_customers=birthday_customers,
        top_customers=top_customers,
        chart_labels=chart_labels,
        chart_data=chart_data,
        is_demo=session.get("is_demo", False),
    )

@app.route("/customers")
def customers():
    b = current_business()
    if not b:
        return redirect("/login")
    customers_list = Customer.query.filter_by(business_id=b.id).order_by(Customer.created_at.desc()).all()
    return render_template("customers.html", customers=customers_list, plan=b.plan)

@app.route("/add-customer", methods=["GET", "POST"])
def add_customer():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        customer_limit = get_plan_limit(b.plan, "customers")
        if customer_limit:
            count = Customer.query.filter_by(business_id=b.id).count()
            if count >= customer_limit:
                flash(f"Customer limit reached ({customer_limit}). Upgrade your plan.", "error")
                return redirect("/upgrade")
        first_name = request.form.get("first_name", "").strip()
        email = request.form.get("email", "").strip() or None
        phone = request.form.get("phone", "").strip()
        if not first_name:
            flash("First name is required.", "error")
            return redirect("/add-customer")
        if not email and not phone:
            flash("Please provide at least a phone number or email.", "error")
            return redirect("/add-customer")
        customer = Customer(
            business_id=b.id,
            first_name=first_name,
            last_name=request.form.get("last_name", "").strip(),
            email=email,
            phone=phone,
            dob=request.form.get("dob", "").strip()
        )
        try:
            db.session.add(customer)
            db.session.commit()
            enroll_customer_in_journeys(customer, trigger="signup")
            flash(f"Customer {first_name} added!", "success")
            return redirect("/customers")
        except:
            db.session.rollback()
            flash("Error adding customer.", "error")
    return render_template("add_customer.html")

@app.route("/edit-customer/<int:customer_id>", methods=["POST"])
def edit_customer(customer_id):
    b = current_business()
    if not b:
        return redirect("/login")
    customer = Customer.query.get(customer_id)
    if not customer or customer.business_id != b.id:
        flash("Customer not found.", "error")
        return redirect("/customers")
    customer.first_name = request.form.get("first_name", customer.first_name).strip() or customer.first_name
    customer.last_name = request.form.get("last_name", "").strip()
    customer.email = request.form.get("email", customer.email).strip() or customer.email
    customer.phone = request.form.get("phone", "").strip()
    customer.dob = request.form.get("dob", "").strip()
    try:
        db.session.commit()
        flash("Customer updated.", "success")
    except:
        db.session.rollback()
        flash("Error updating customer.", "error")
    return redirect("/customers")

@app.route("/delete-customer/<int:customer_id>", methods=["POST"])
def delete_customer(customer_id):
    b = current_business()
    if not b:
        return redirect("/login")
    customer = Customer.query.get(customer_id)
    if customer and customer.business_id == b.id:
        try:
            db.session.delete(customer)
            db.session.commit()
            flash("Customer deleted.", "success")
        except:
            db.session.rollback()
            flash("Error deleting customer.", "error")
    return redirect("/customers")

@app.route("/campaigns")
def campaigns():
    b = current_business()
    if not b:
        return redirect("/login")
    page = request.args.get("page", 1, type=int)
    per_page = 20
    pagination = Campaign.query.filter_by(business_id=b.id).order_by(Campaign.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template("campaigns.html", campaigns=pagination.items, pagination=pagination, campaign_types=get_campaign_types())

@app.route("/create-campaign", methods=["GET", "POST"])
def create_campaign():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        customer_name = request.form.get("customer_name", "").strip()
        customer_email = request.form.get("customer_email", "").strip()
        customer_phone = request.form.get("customer_phone", "").strip()
        campaign_type = request.form.get("campaign_type", "promotion").strip()
        use_ai = request.form.get("use_ai") == "on"
        message = request.form.get("message", "").strip()
        scheduled_at_str = request.form.get("scheduled_at", "").strip()
        if not customer_name:
            flash("Customer name is required.", "error")
            return redirect("/create-campaign")
        if not customer_email and not customer_phone:
            flash("Please provide at least an email or phone number.", "error")
            return redirect("/create-campaign")
        if use_ai and not message:
            message = generate_ai_message(customer_name, b.business_name, campaign_type)
        elif not message:
            flash("Provide message or use AI.", "error")
            return redirect("/create-campaign")
        scheduled_at = None
        if scheduled_at_str:
            try:
                scheduled_at = datetime.strptime(scheduled_at_str, "%Y-%m-%dT%H:%M")
                if scheduled_at <= datetime.utcnow():
                    flash("Scheduled time must be in the future.", "error")
                    return redirect("/create-campaign")
            except ValueError:
                flash("Invalid schedule date. Please use the date picker.", "error")
                return redirect("/create-campaign")
        campaign_status = "scheduled" if scheduled_at else "draft"
        campaign = Campaign(
            business_id=b.id,
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            campaign_type=campaign_type,
            message=message,
            status=campaign_status,
            scheduled_at=scheduled_at
        )
        try:
            db.session.add(campaign)
            db.session.commit()
            if scheduled_at:
                flash(f"Campaign scheduled for {scheduled_at.strftime('%b %d at %I:%M %p')}!", "success")
            else:
                flash("Campaign created!", "success")
            return redirect(f"/campaign/{campaign.id}")
        except:
            db.session.rollback()
            flash("Error creating campaign.", "error")
    customers_list = Customer.query.filter_by(business_id=b.id).all()
    # Pre-load a saved promotion if ?promo_id= passed
    promo_id = request.args.get("promo_id")
    selected_promo = None
    if promo_id:
        p = Promotion.query.get(int(promo_id))
        if p and p.business_id == b.id:
            selected_promo = {"id": p.id, "name": p.name, "description": p.description, "price": p.price, "category": p.category}
    return render_template("create_campaign.html", campaign_types=get_campaign_types(), customers=customers_list, selected_promo=selected_promo)

@app.route("/campaign/<int:campaign_id>")
def view_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    campaign = Campaign.query.get(campaign_id)
    if not campaign or campaign.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    return render_template("view_campaign.html", campaign=campaign, plan=b.plan, campaign_types=get_campaign_types())

@app.route("/unsubscribe/<token>")
def unsubscribe(token):
    campaign_id = verify_unsubscribe_token(token)
    if not campaign_id:
        return render_template("404.html"), 404
    campaign = Campaign.query.get(campaign_id)
    if campaign:
        # Mark all customers with this email (for this business) as unsubscribed
        Customer.query.filter_by(
            business_id=campaign.business_id,
            email=campaign.customer_email
        ).update({"unsubscribed": True})
        db.session.commit()
    return render_template("unsubscribed.html")

@app.route("/track/open/<int:campaign_id>")
def track_open(campaign_id):
    campaign = Campaign.query.get(campaign_id)
    if campaign:
        campaign.open_count = (campaign.open_count or 0) + 1
        if not campaign.opened_at:
            campaign.opened_at = datetime.utcnow()
        try:
            db.session.commit()
        except:
            db.session.rollback()
    # Return a 1x1 transparent GIF
    pixel = b'\x47\x49\x46\x38\x39\x61\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00\x21\xf9\x04\x00\x00\x00\x00\x00\x2c\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02\x44\x01\x00\x3b'
    from flask import Response
    return Response(pixel, mimetype="image/gif", headers={"Cache-Control": "no-store, no-cache, must-revalidate"})

@app.route("/track/click/<int:campaign_id>")
def track_click(campaign_id):
    campaign = Campaign.query.get(campaign_id)
    destination = "/"
    if campaign:
        campaign.click_count = (campaign.click_count or 0) + 1
        if not campaign.clicked_at:
            campaign.clicked_at = datetime.utcnow()
        try:
            db.session.commit()
        except:
            db.session.rollback()
        b = Business.query.get(campaign.business_id)
        if b and b.website:
            destination = b.website if b.website.startswith("http") else f"https://{b.website}"
        elif b and b.address:
            destination = f"https://www.google.com/maps/search/?api=1&query={b.address.replace(' ', '+')}"
    return redirect(destination)

@app.route("/send-campaign/<int:campaign_id>", methods=["POST"])
def send_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    campaign = Campaign.query.get(campaign_id)
    if not campaign or campaign.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    if session.get("is_demo"):
        flash("Demo mode — emails are not sent. Sign up for a free account to send real campaigns!", "success")
        return redirect(f"/campaign/{campaign.id}")
    # Check if customer is unsubscribed
    customer = Customer.query.filter_by(business_id=b.id, email=campaign.customer_email).first()
    if customer and customer.unsubscribed:
        flash("This customer has unsubscribed and cannot receive emails.", "error")
        return redirect(f"/campaign/{campaign.id}")
    token = get_unsubscribe_token(campaign.id)
    unsub_url = url_for("unsubscribe", token=token, _external=True)
    pixel_url = url_for("track_open", campaign_id=campaign.id, _external=True)
    click_url = url_for("track_click", campaign_id=campaign.id, _external=True)
    subject = f"Special Offer from {b.business_name}"
    success = send_email(
        campaign.customer_email, subject, campaign.message,
        customer_name=campaign.customer_name,
        business_name=b.business_name,
        campaign_type=campaign.campaign_type,
        unsubscribe_url=unsub_url,
        business_address=b.address or "",
        business_phone=b.phone or "",
        business_website=b.website or "",
        tracking_pixel_url=pixel_url,
        click_tracking_url=click_url,
        business_reply_email=b.email
    )
    if success:
        campaign.status = "sent"
        try:
            db.session.commit()
            flash("Campaign sent successfully!", "success")
        except:
            db.session.rollback()
            flash("Error updating campaign.", "error")
    else:
        flash("Email not sent. Check email configuration.", "error")
    return redirect(f"/campaign/{campaign.id}")

@app.route("/send-sms/<int:campaign_id>", methods=["POST"])
def send_sms_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    if session.get("is_demo"):
        flash("Demo mode — SMS not sent. Sign up for a free account to send real campaigns!", "success")
        return redirect(f"/campaign/{campaign_id}")
    if b.plan == "free":
        flash("SMS requires Starter or Pro plan.", "error")
        return redirect("/upgrade")
    campaign = Campaign.query.get(campaign_id)
    if not campaign or campaign.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    if not campaign.customer_phone:
        flash("No phone number for this customer.", "error")
        return redirect(f"/campaign/{campaign.id}")
    success = send_sms(campaign.customer_phone, campaign.message)
    if success:
        flash("SMS sent!", "success")
    else:
        flash("SMS not sent. Check Twilio configuration in Render environment.", "error")
    return redirect(f"/campaign/{campaign.id}")

@app.route("/quick-sms", methods=["GET", "POST"])
def quick_sms():
    b = current_business()
    if not b:
        return redirect("/login")
    if b.plan == "free":
        flash("Quick SMS requires Starter or Pro plan.", "error")
        return redirect("/upgrade")
    if request.method == "POST":
        recipient_name = request.form.get("recipient_name", "").strip() or "Customer"
        phone = request.form.get("phone", "").strip()
        message = request.form.get("message", "").strip()
        use_ai = request.form.get("use_ai") == "on"
        campaign_type = request.form.get("campaign_type", "promotion")
        if not phone:
            flash("Phone number is required.", "error")
            return redirect("/quick-sms")
        if use_ai and not message:
            message = generate_ai_message(recipient_name, b.business_name, campaign_type)
        if not message:
            flash("Enter a message or enable AI generation.", "error")
            return redirect("/quick-sms")
        success = send_sms(phone, message)
        # Save as a campaign record for tracking
        campaign = Campaign(
            business_id=b.id,
            customer_name=recipient_name,
            customer_email="",
            customer_phone=phone,
            campaign_type=campaign_type,
            message=message,
            status="sent" if success else "failed",
        )
        try:
            db.session.add(campaign)
            db.session.commit()
        except Exception:
            db.session.rollback()
        if success:
            flash(f"SMS sent to {phone}!", "success")
        else:
            flash("SMS not sent. Check Twilio configuration.", "error")
        return redirect("/quick-sms")
    customers_list = Customer.query.filter_by(business_id=b.id).filter(Customer.phone != None).filter(Customer.phone != "").order_by(Customer.first_name).all()
    return render_template("quick_sms.html", campaign_types=get_campaign_types(), customers=customers_list, plan=b.plan)

@app.route("/bulk-send", methods=["GET", "POST"])
def bulk_send():
    b = current_business()
    if not b:
        return redirect("/login")
    if session.get("is_demo") and request.method == "POST":
        flash("Demo mode — emails not sent. Sign up for a free account to send real campaigns!", "success")
        return redirect("/bulk-send")
    if request.method == "POST":
        campaign_type = request.form.get("campaign_type", "promotion")
        use_ai = request.form.get("use_ai") == "on"
        custom_message = request.form.get("message", "").strip()
        scheduled_at_str = request.form.get("scheduled_at", "").strip()
        segment = request.form.get("segment", "all")
        customers_list = get_segment_customers(b.id, segment)
        if not customers_list:
            flash("No customers match that segment. Try a different one.", "error")
            return redirect("/bulk-send")
        if not custom_message and not use_ai:
            flash("Enter a message or enable AI generation.", "error")
            return redirect("/bulk-send")
        scheduled_at = None
        if scheduled_at_str:
            try:
                scheduled_at = datetime.strptime(scheduled_at_str, "%Y-%m-%dT%H:%M")
            except ValueError:
                pass
        sent_count = 0
        skipped_unsub = 0
        skipped_no_email = 0
        for customer in customers_list:
            if customer.unsubscribed:
                skipped_unsub += 1
                continue
            if not customer.email:
                skipped_no_email += 1
                continue
            msg = generate_ai_message(customer.first_name, b.business_name, campaign_type) if (use_ai and not custom_message) else custom_message
            campaign_status = "scheduled" if scheduled_at else "draft"
            campaign = Campaign(
                business_id=b.id,
                customer_name=customer.first_name,
                customer_email=customer.email,
                customer_phone=customer.phone,
                campaign_type=campaign_type,
                message=msg,
                status=campaign_status,
                scheduled_at=scheduled_at
            )
            db.session.add(campaign)
            if not scheduled_at:
                db.session.flush()
                token = get_unsubscribe_token(campaign.id)
                unsub_url = url_for("unsubscribe", token=token, _external=True)
                pixel_url = url_for("track_open", campaign_id=campaign.id, _external=True)
                click_url = url_for("track_click", campaign_id=campaign.id, _external=True)
                subject = f"Special Offer from {b.business_name}"
                if send_email(customer.email, subject, msg, customer_name=customer.first_name, business_name=b.business_name, campaign_type=campaign_type, unsubscribe_url=unsub_url, business_address=b.address or "", business_phone=b.phone or "", business_website=b.website or "", tracking_pixel_url=pixel_url, click_tracking_url=click_url, business_reply_email=b.email):
                    campaign.status = "sent"
                    sent_count += 1
        db.session.commit()
        notes = []
        if skipped_unsub: notes.append(f"{skipped_unsub} unsubscribed")
        if skipped_no_email: notes.append(f"{skipped_no_email} phone-only (no email)")
        note_str = f" ({', '.join(notes)} skipped)" if notes else ""
        if scheduled_at:
            flash(f"Scheduled {len(customers_list) - skipped_unsub - skipped_no_email} campaigns for {scheduled_at.strftime('%b %d at %I:%M %p')} UTC{note_str}.", "success")
        else:
            flash(f"Bulk send complete! Sent to {sent_count}/{len(customers_list)} customers{note_str}.", "success")
        return redirect("/campaigns")
    customers_count = Customer.query.filter_by(business_id=b.id, unsubscribed=False).count()
    preselect_segment = request.args.get("segment", "all")
    return render_template("bulk_send.html", campaign_types=get_campaign_types(), customers_count=customers_count, preselect_segment=preselect_segment)

@app.route("/edit-campaign/<int:campaign_id>", methods=["POST"])
def edit_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    campaign = Campaign.query.get(campaign_id)
    if not campaign or campaign.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    if campaign.status == "sent":
        flash("Cannot edit a sent campaign.", "error")
        return redirect(f"/campaign/{campaign.id}")
    campaign.message = request.form.get("message", campaign.message).strip() or campaign.message
    campaign.campaign_type = request.form.get("campaign_type", campaign.campaign_type)
    scheduled_at_str = request.form.get("scheduled_at", "").strip()
    if scheduled_at_str:
        try:
            campaign.scheduled_at = datetime.strptime(scheduled_at_str, "%Y-%m-%dT%H:%M")
            campaign.status = "scheduled"
        except ValueError:
            pass
    elif campaign.status == "scheduled":
        campaign.status = "draft"
        campaign.scheduled_at = None
    try:
        db.session.commit()
        flash("Campaign updated.", "success")
    except:
        db.session.rollback()
        flash("Error updating campaign.", "error")
    return redirect(f"/campaign/{campaign.id}")

@app.route("/clone-campaign/<int:campaign_id>", methods=["POST"])
def clone_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    original = Campaign.query.get(campaign_id)
    if not original or original.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    clone = Campaign(
        business_id=b.id,
        customer_name=original.customer_name,
        customer_email=original.customer_email,
        customer_phone=original.customer_phone,
        campaign_type=original.campaign_type,
        message=original.message,
        status="draft"
    )
    try:
        db.session.add(clone)
        db.session.commit()
        flash("Campaign cloned as a new draft.", "success")
        return redirect(f"/campaign/{clone.id}")
    except:
        db.session.rollback()
        flash("Error cloning campaign.", "error")
        return redirect(f"/campaign/{campaign_id}")

@app.route("/test-email/<int:campaign_id>", methods=["POST"])
def test_email(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    if session.get("is_demo"):
        flash("Demo mode — test email not sent. Sign up for a free account!", "success")
        return redirect(f"/campaign/{campaign_id}")
    campaign = Campaign.query.get(campaign_id)
    if not campaign or campaign.business_id != b.id:
        flash("Campaign not found.", "error")
        return redirect("/campaigns")
    subject = f"[TEST] {b.business_name} — {get_campaign_types().get(campaign.campaign_type, 'Campaign')}"
    success = send_email(
        b.email, subject, campaign.message,
        customer_name=b.owner_name,
        business_name=b.business_name,
        campaign_type=campaign.campaign_type,
        business_address=b.address or "",
        business_phone=b.phone or "",
        business_website=b.website or "",
        business_reply_email=b.email
    )
    if success:
        flash(f"Test email sent to {b.email}!", "success")
    else:
        flash("Test email failed. Check email configuration.", "error")
    return redirect(f"/campaign/{campaign.id}")

@app.route("/delete-campaign/<int:campaign_id>", methods=["POST"])
def delete_campaign(campaign_id):
    b = current_business()
    if not b:
        return redirect("/login")
    campaign = Campaign.query.get(campaign_id)
    if campaign and campaign.business_id == b.id:
        try:
            db.session.delete(campaign)
            db.session.commit()
            flash("Campaign deleted.", "success")
        except:
            db.session.rollback()
            flash("Error deleting campaign.", "error")
    return redirect("/campaigns")

@app.route("/upload-customers", methods=["GET", "POST"])
def upload_customers():
    b = current_business()
    if not b:
        return redirect("/login")

    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Please select a file.", "error")
            return redirect("/upload-customers")

        filename = file.filename.lower()
        rows = []

        try:
            if filename.endswith(".csv"):
                import csv, io
                content = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row)})
            else:
                flash("Only CSV or Excel (.xlsx) files supported.", "error")
                return redirect("/upload-customers")
        except Exception as e:
            flash(f"Error reading file: {e}", "error")
            return redirect("/upload-customers")

        # Auto-detect columns
        def find_col(row, keywords):
            for key in row.keys():
                if any(kw in key.lower() for kw in keywords):
                    return row.get(key, "").strip()
            return ""

        added = 0
        skipped = 0
        for row in rows:
            email = find_col(row, ["email", "e-mail", "mail"])
            phone = find_col(row, ["phone", "mobile", "cell", "tel", "number"])
            first_name = find_col(row, ["first", "fname", "name"]) or "Customer"
            last_name = find_col(row, ["last", "lname", "surname"])

            if not email and not phone:
                skipped += 1
                continue

            # Check plan limits
            customer_limit = get_plan_limit(b.plan, "customers")
            if customer_limit:
                count = Customer.query.filter_by(business_id=b.id).count()
                if count >= customer_limit:
                    flash(f"Customer limit reached ({customer_limit}). Upgrade to import more.", "error")
                    break

            customer = Customer(
                business_id=b.id,
                first_name=first_name,
                last_name=last_name,
                email=email or None,
                phone=phone or None,
            )
            db.session.add(customer)
            added += 1

        try:
            db.session.commit()
            flash(f"Imported {added} customers! ({skipped} rows skipped — no email or phone found)", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving customers: {e}", "error")

        return redirect("/customers")

    return render_template("upload_customers.html")

@app.route("/paste-customers", methods=["POST"])
def paste_customers():
    b = current_business()
    if not b:
        return redirect("/login")

    raw = request.form.get("contacts", "").strip()
    default_name = request.form.get("default_name", "Customer").strip() or "Customer"

    if not raw:
        flash("Please paste some contacts.", "error")
        return redirect("/upload-customers")

    import re
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    added = 0

    for line in lines:
        # Check plan limit
        customer_limit = get_plan_limit(b.plan, "customers")
        if customer_limit:
            count = Customer.query.filter_by(business_id=b.id).count()
            if count >= customer_limit:
                flash(f"Customer limit reached ({customer_limit}). Upgrade to import more.", "error")
                break

        email = ""
        phone = ""

        # Detect if line is email
        if "@" in line:
            email = line
        else:
            # Clean and use as phone
            phone = re.sub(r"[^\d\+\-\(\)\s]", "", line).strip()
            if len(re.sub(r"\D", "", phone)) < 7:
                continue  # skip if less than 7 digits

        customer = Customer(
            business_id=b.id,
            first_name=default_name,
            email=email or None,
            phone=phone or None,
        )
        db.session.add(customer)
        added += 1

    try:
        db.session.commit()
        flash(f"Imported {added} contacts successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error saving contacts: {e}", "error")

    return redirect("/customers")

@app.route("/generate-message", methods=["POST"])
def generate_message():
    b = current_business()
    if not b:
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    try:
        data = request.get_json()
        customer_name = data.get("customer_name", "").strip()
        campaign_type = data.get("campaign_type", "promotion").strip()
        if not customer_name:
            return jsonify({"success": False, "error": "Customer name required"})
        try:
            message = generate_ai_message(customer_name, b.business_name, campaign_type, raise_on_error=True)
            return jsonify({"success": True, "message": message, "ai": True})
        except Exception as ai_err:
            print(f"AI error, using fallback: {ai_err}")
            message = generate_ai_message(customer_name, b.business_name, campaign_type, raise_on_error=False)
            return jsonify({"success": True, "message": message, "ai": False, "warning": "AI unavailable — using a template message. Feel free to edit it."})
    except Exception as e:
        print(f"Generate message error: {e}")
        return jsonify({"success": False, "error": f"Error: {str(e)}"})

@app.route("/promotions", methods=["GET", "POST"])
def promotions():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        price = request.form.get("price", "").strip()
        category = request.form.get("category", "promotion").strip()
        if not name:
            flash("Name is required.", "error")
            return redirect("/promotions")
        db.session.add(Promotion(
            business_id=b.id, name=name, description=description,
            price=price, category=category
        ))
        db.session.commit()
        flash(f"'{name}' saved to your promotions library!", "success")
        return redirect("/promotions")
    promos = Promotion.query.filter_by(business_id=b.id, active=True).order_by(Promotion.created_at.desc()).all()
    return render_template("promotions.html", promotions=promos, campaign_types=get_campaign_types())

@app.route("/promotions/delete/<int:promo_id>", methods=["POST"])
def delete_promotion(promo_id):
    b = current_business()
    if not b:
        return redirect("/login")
    p = Promotion.query.get(promo_id)
    if p and p.business_id == b.id:
        db.session.delete(p)
        db.session.commit()
        flash("Promotion deleted.", "success")
    return redirect("/promotions")

@app.route("/promotions/edit/<int:promo_id>", methods=["POST"])
def edit_promotion(promo_id):
    b = current_business()
    if not b:
        return redirect("/login")
    p = Promotion.query.get(promo_id)
    if not p or p.business_id != b.id:
        flash("Not found.", "error")
        return redirect("/promotions")
    p.name = request.form.get("name", p.name).strip() or p.name
    p.description = request.form.get("description", "").strip()
    p.price = request.form.get("price", "").strip()
    p.category = request.form.get("category", p.category)
    db.session.commit()
    flash("Updated.", "success")
    return redirect("/promotions")

@app.route("/api/promotions")
def api_promotions():
    b = current_business()
    if not b:
        return jsonify([])
    promos = Promotion.query.filter_by(business_id=b.id, active=True).order_by(Promotion.created_at.desc()).all()
    return jsonify([{"id": p.id, "name": p.name, "description": p.description, "price": p.price, "category": p.category} for p in promos])

AUTOMATION_DEFAULTS = {
    "birthday": {
        "label": "Birthday Campaign",
        "icon": "🎂",
        "desc": "Automatically sends a birthday message to guests on (or before) their birthday.",
        "default_msg": "Happy Birthday {name}! 🎉 As our special guest, enjoy a complimentary gift on your next visit to {business}. You deserve to be celebrated!",
        "default_days": 0,
        "days_label": "Days before birthday to send",
    },
    "reengagement": {
        "label": "Re-engagement Campaign",
        "icon": "🔄",
        "desc": "Automatically reaches out to guests you haven't contacted in a while.",
        "default_msg": "Hi {name}, we miss you at {business}! 😊 It's been a while — come back and enjoy an exclusive offer just for you. We'd love to see you again!",
        "default_days": 30,
        "days_label": "Days since last contact",
    },
    "welcome": {
        "label": "Welcome Message",
        "icon": "👋",
        "desc": "Automatically sends a warm welcome to every new guest added to your list.",
        "default_msg": "Welcome to {business}, {name}! 🎉 We're thrilled to have you. Stay tuned for exclusive offers and updates just for our valued guests.",
        "default_days": 0,
        "days_label": "",
    },
}

@app.route("/automations", methods=["GET", "POST"])
def automations():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        rule_type = request.form.get("rule_type")
        if rule_type not in AUTOMATION_DEFAULTS:
            flash("Invalid automation type.", "error")
            return redirect("/automations")
        active = request.form.get("active") == "on"
        message_template = request.form.get("message_template", "").strip()
        days_threshold = int(request.form.get("days_threshold", 0) or 0)
        rule = AutomationRule.query.filter_by(business_id=b.id, rule_type=rule_type).first()
        if rule:
            rule.active = active
            rule.message_template = message_template
            rule.days_threshold = days_threshold
        else:
            rule = AutomationRule(
                business_id=b.id, rule_type=rule_type, active=active,
                message_template=message_template, days_threshold=days_threshold
            )
            db.session.add(rule)
        db.session.commit()
        status = "enabled" if active else "disabled"
        flash(f"{AUTOMATION_DEFAULTS[rule_type]['label']} {status}.", "success")
        return redirect("/automations")
    rules = {r.rule_type: r for r in AutomationRule.query.filter_by(business_id=b.id).all()}
    return render_template("automations.html", rules=rules, defaults=AUTOMATION_DEFAULTS)

@app.route("/api/segment-count")
def segment_count():
    b = current_business()
    if not b:
        return jsonify({"count": 0})
    segment = request.args.get("segment", "all")
    customers = get_segment_customers(b.id, segment)
    sample = [{"name": f"{c.first_name} {c.last_name or ''}".strip(), "email": c.email or "", "phone": c.phone or ""} for c in customers[:5]]
    return jsonify({"count": len(customers), "sample": sample})

@app.route("/ai-ideas", methods=["POST"])
def ai_ideas():
    b = current_business()
    if not b:
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    try:
        data = request.get_json()
        description = data.get("description", "").strip()
        customer_name = data.get("customer_name", "a valued customer").strip() or "a valued customer"
        campaign_type = data.get("campaign_type", "promotion").strip()
        if not description:
            return jsonify({"success": False, "error": "Description required"})
        if not client:
            return jsonify({"success": False, "error": "AI not configured"})
        prompt = (
            f"You are a marketing expert for a small business called '{b.business_name}'.\n"
            f"The business owner wants to promote this: \"{description}\"\n"
            f"The customer's name is {customer_name}.\n"
            f"Write 3 different short SMS/email marketing messages (each under 3 sentences, max 160 chars, with emojis, no markdown).\n"
            f"Return ONLY a JSON array of 3 strings, no explanation. Example: [\"msg1\", \"msg2\", \"msg3\"]"
        )
        import json as json_module
        raw = clean_ai_text(_call_openai(prompt, max_tokens=400))
        # Extract JSON array from response
        start = raw.find("[")
        end = raw.rfind("]") + 1
        ideas = json_module.loads(raw[start:end]) if start >= 0 else [raw]
        return jsonify({"success": True, "ideas": ideas[:3]})
    except Exception as e:
        print(f"AI ideas error: {e}")
        return jsonify({"success": False, "error": str(e)})

@app.route("/load-demo", methods=["POST"])
def load_demo():
    b = current_business()
    if not b:
        return redirect("/login")

    # Birthdays: some within the next 7 days (relative to today) so Birthday Radar shows live
    _today = datetime.utcnow()
    def _bday(month, day, year=1988):
        return f"{year}-{month:02d}-{day:02d}"
    def _soon(days_ahead, year=1990):
        d = _today + timedelta(days=days_ahead)
        return f"{year}-{d.month:02d}-{d.day:02d}"

    DEMO_CUSTOMERS = [
        {"first": "James",    "last": "Martinez",  "email": "james.martinez@demo.com",   "phone": "+12145550101", "dob": _soon(1)},
        {"first": "Priya",    "last": "Sharma",    "email": "priya.sharma@demo.com",     "phone": "+12145550102", "dob": _soon(3)},
        {"first": "Carlos",   "last": "Reyes",     "email": "carlos.reyes@demo.com",     "phone": "+12145550103", "dob": _bday(4, 12)},
        {"first": "Ashley",   "last": "Thompson",  "email": "ashley.t@demo.com",         "phone": "+12145550104", "dob": _bday(6, 22)},
        {"first": "Michael",  "last": "Chen",      "email": "michael.chen@demo.com",     "phone": "+12145550105", "dob": _bday(8, 5)},
        {"first": "Fatima",   "last": "Al-Hassan", "email": "fatima.h@demo.com",         "phone": "+12145550106", "dob": _bday(9, 18)},
        {"first": "David",    "last": "Williams",  "email": "david.w@demo.com",          "phone": "+12145550107", "dob": _bday(11, 30)},
        {"first": "Sofia",    "last": "Nguyen",    "email": "sofia.nguyen@demo.com",     "phone": "+12145550108", "dob": _soon(5)},
        {"first": "Kevin",    "last": "Johnson",   "email": "kevin.j@demo.com",          "phone": "+12145550109", "dob": _bday(2, 14)},
        {"first": "Maria",    "last": "Garcia",    "email": "maria.garcia@demo.com",     "phone": "+12145550110", "dob": _bday(7, 4)},
        {"first": "Tyler",    "last": "Brooks",    "email": "tyler.brooks@demo.com",     "phone": "+12145550111", "dob": _bday(10, 31)},
        {"first": "Aisha",    "last": "Patel",     "email": "aisha.patel@demo.com",      "phone": "+12145550112", "dob": _soon(2)},
        {"first": "Ryan",     "last": "Kim",       "email": "ryan.kim@demo.com",         "phone": "+12145550113", "dob": _bday(5, 20)},
        {"first": "Jessica",  "last": "Davis",     "email": "jessica.d@demo.com",        "phone": "+12145550114", "dob": _bday(12, 1)},
        {"first": "Brandon",  "last": "Lee",       "email": "brandon.lee@demo.com",      "phone": "+12145550115", "dob": _bday(3, 8)},
        {"first": "Natalie",  "last": "Robinson",  "email": "natalie.r@demo.com",        "phone": "+12145550116", "dob": _bday(1, 25)},
        {"first": "Omar",     "last": "Hassan",    "email": "omar.hassan@demo.com",      "phone": "+12145550117", "dob": _bday(6, 15)},
        {"first": "Lauren",   "last": "Mitchell",  "email": "lauren.m@demo.com",         "phone": "+12145550118", "dob": _bday(4, 3)},
        {"first": "Ethan",    "last": "Cooper",    "email": "ethan.c@demo.com",          "phone": "+12145550119", "dob": _bday(8, 27)},
        {"first": "Rachel",   "last": "Torres",    "email": "rachel.t@demo.com",         "phone": "+12145550120", "dob": _bday(11, 11)},
    ]

    DEMO_CAMPAIGNS = [
        {"name": "James Martinez",   "email": "james.martinez@demo.com",  "phone": "+12145550101", "type": "come_back",  "status": "sent",
         "msg": "Hey James! We miss you at {}! Come back this week and enjoy 15% off your next visit. Use code: COMEBACK15 — we'd love to see you again!"},
        {"name": "Priya Sharma",     "email": "priya.sharma@demo.com",    "phone": "+12145550102", "type": "birthday",   "status": "sent",
         "msg": "Happy Birthday Priya! From all of us at {}, we hope your day is amazing! Enjoy 20% off your next visit — our gift to you. Use code: BDAY20"},
        {"name": "Carlos Reyes",     "email": "carlos.reyes@demo.com",    "phone": "+12145550103", "type": "weekend",    "status": "sent",
         "msg": "Carlos, this weekend only — {} is running an exclusive special just for our regulars! Come in Saturday or Sunday for amazing deals. See you soon!"},
        {"name": "Ashley Thompson",  "email": "ashley.t@demo.com",        "phone": "+12145550104", "type": "loyalty",    "status": "sent",
         "msg": "Ashley, you are one of our most valued customers at {}! As a thank-you for your loyalty, here's a special reward waiting for you — ask us when you visit!"},
        {"name": "Michael Chen",     "email": "michael.chen@demo.com",    "phone": "+12145550105", "type": "lunch",      "status": "sent",
         "msg": "Michael, join us for lunch at {}! This week we have incredible daily specials, fresh ingredients, and your favorite dishes ready to go. Come see us!"},
        {"name": "Fatima Al-Hassan", "email": "fatima.h@demo.com",        "phone": "+12145550106", "type": "new_item",   "status": "sent",
         "msg": "Fatima, exciting news! {} just launched something new and we think you're going to love it. Come in and be one of the first to try it!"},
        {"name": "David Williams",   "email": "david.w@demo.com",         "phone": "+12145550107", "type": "happy_hour", "status": "sent",
         "msg": "David! Happy Hour at {} is happening this week — amazing drinks, great bites, and unbeatable prices. Bring a friend and make it a night!"},
        {"name": "Sofia Nguyen",     "email": "sofia.nguyen@demo.com",    "phone": "+12145550108", "type": "dinner",     "status": "sent",
         "msg": "Sofia, treat yourself tonight! {} has a special dinner offer just for you — a memorable evening with exceptional food. Reserve your spot today!"},
        {"name": "Kevin Johnson",    "email": "kevin.j@demo.com",         "phone": "+12145550109", "type": "promotion",  "status": "sent",
         "msg": "Kevin, {} has an exclusive promotion running this week just for our VIP customers! Don't miss out — visit us and mention this message at the door."},
        {"name": "Maria Garcia",     "email": "maria.garcia@demo.com",    "phone": "+12145550110", "type": "come_back",  "status": "draft",
         "msg": "Maria, it's been a while and we miss you! Come back to {} this week and we'll make sure you feel like a VIP. See you soon!"},
        {"name": "Tyler Brooks",     "email": "tyler.brooks@demo.com",    "phone": "+12145550111", "type": "weekend",    "status": "draft",
         "msg": "Tyler! Big weekend coming up at {}. We have something special planned and want YOU there. Saturday and Sunday only — come check it out!"},
    ]

    # Clear existing demo data first (customers with @demo.com emails)
    existing_emails = [c["email"] for c in DEMO_CUSTOMERS]
    Customer.query.filter(
        Customer.business_id == b.id,
        Customer.email.in_(existing_emails)
    ).delete(synchronize_session=False)
    Campaign.query.filter(
        Campaign.business_id == b.id,
        Campaign.customer_email.in_(existing_emails)
    ).delete(synchronize_session=False)

    # Add demo customers
    for c in DEMO_CUSTOMERS:
        db.session.add(Customer(
            business_id=b.id,
            first_name=c["first"],
            last_name=c["last"],
            email=c["email"],
            phone=c["phone"],
            dob=c.get("dob"),
        ))

    # Add demo campaigns
    for c in DEMO_CAMPAIGNS:
        db.session.add(Campaign(
            business_id=b.id,
            customer_name=c["name"],
            customer_email=c["email"],
            customer_phone=c["phone"],
            campaign_type=c["type"],
            message=c["msg"].format(b.business_name),
            status=c["status"],
        ))

    try:
        db.session.commit()
        flash(f"Demo data loaded! 20 customers and {len(DEMO_CAMPAIGNS)} campaigns added. You're ready to record.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error loading demo data: {e}", "error")

    return redirect("/dashboard")

DEMO_EMAIL = "demo@revvio.ai"

@app.route("/demo-login")
def demo_login():
    """One-click demo: creates/resets the demo account and logs in."""
    # Find or create demo business
    demo = Business.query.filter_by(email=DEMO_EMAIL).first()
    if not demo:
        hashed = bcrypt.hashpw(b"demo-revvio-2024", bcrypt.gensalt()).decode("utf-8")
        demo = Business(
            business_name="Mario's Grill",
            owner_name="Demo User",
            email=DEMO_EMAIL,
            password=hashed,
            plan="starter",
            address="123 Main St, Dallas, TX 75201"
        )
        db.session.add(demo)
        db.session.flush()  # get demo.id before commit

    # Ensure demo has a completed profile
    profile = BusinessProfile.query.filter_by(business_id=demo.id).first()
    if not profile:
        profile = BusinessProfile(
            business_id=demo.id,
            cuisine_type="Italian-American",
            signature_dish="Wood-fired Pizza",
            special_offer="15% off first visit",
            slow_days="Monday,Tuesday",
            peak_days="Friday,Saturday,Sunday",
            tone="friendly",
            timezone="America/Chicago",
            auto_welcome=True,
            auto_weekly=True,
            auto_flash=True,
            auto_birthday=True,
            auto_winback=True,
            auto_review=True,
            auto_loyalty=True,
            weekly_send_day="Tuesday",
            setup_complete=True,
        )
        db.session.add(profile)
    else:
        profile.setup_complete = True

    # Wipe and re-seed demo data
    Customer.query.filter(Customer.business_id == demo.id).delete(synchronize_session=False)
    Campaign.query.filter(Campaign.business_id == demo.id).delete(synchronize_session=False)
    # Seed demo journeys
    existing_journeys = Journey.query.filter_by(business_id=demo.id).all()
    for j in existing_journeys:
        JourneyStep.query.filter_by(journey_id=j.id).delete(synchronize_session=False)
        CustomerJourney.query.filter_by(journey_id=j.id).delete(synchronize_session=False)
    Journey.query.filter_by(business_id=demo.id).delete(synchronize_session=False)
    demo_journeys = [
        {"name": "New Customer Welcome Journey", "trigger": "signup", "steps": [
            {"order": 0, "delay": 0,  "type": "loyalty",   "use_ai": True, "channel": "email"},
            {"order": 1, "delay": 3,  "type": "promotion", "use_ai": True, "channel": "email"},
            {"order": 2, "delay": 7,  "type": "come_back", "use_ai": True, "channel": "email", "condition": "not_visited"},
            {"order": 3, "delay": 14, "type": "weekend",   "use_ai": True, "channel": "email"},
        ]},
        {"name": "Win-Back Journey", "trigger": "winback", "steps": [
            {"order": 0, "delay": 0,  "type": "come_back", "use_ai": True, "channel": "email"},
            {"order": 1, "delay": 7,  "type": "promotion", "use_ai": True, "channel": "email"},
            {"order": 2, "delay": 14, "type": "come_back", "use_ai": True, "channel": "sms"},
        ]},
        {"name": "Birthday VIP Journey", "trigger": "birthday", "steps": [
            {"order": 0, "delay": 0, "type": "birthday",   "use_ai": True, "channel": "email"},
            {"order": 1, "delay": 3, "type": "come_back",  "use_ai": True, "channel": "email", "condition": "not_visited"},
        ]},
    ]
    for jd in demo_journeys:
        j = Journey(business_id=demo.id, name=jd["name"], trigger=jd["trigger"], active=True)
        db.session.add(j)
        db.session.flush()
        for s in jd["steps"]:
            db.session.add(JourneyStep(
                journey_id=j.id, step_order=s["order"], delay_days=s["delay"],
                message_type=s["type"], use_ai=s["use_ai"], channel=s["channel"],
                condition=s.get("condition", "none")
            ))

    DEMO_CUSTOMERS = [
        {"first": "James",   "last": "Martinez",  "email": "james.martinez@demo.com",  "phone": "+12145550101"},
        {"first": "Priya",   "last": "Sharma",    "email": "priya.sharma@demo.com",    "phone": "+12145550102"},
        {"first": "Carlos",  "last": "Reyes",     "email": "carlos.reyes@demo.com",    "phone": "+12145550103"},
        {"first": "Ashley",  "last": "Thompson",  "email": "ashley.t@demo.com",        "phone": "+12145550104"},
        {"first": "Michael", "last": "Chen",      "email": "michael.chen@demo.com",    "phone": "+12145550105"},
        {"first": "Fatima",  "last": "Al-Hassan", "email": "fatima.h@demo.com",        "phone": "+12145550106"},
        {"first": "David",   "last": "Williams",  "email": "david.w@demo.com",         "phone": "+12145550107"},
        {"first": "Sofia",   "last": "Nguyen",    "email": "sofia.nguyen@demo.com",    "phone": "+12145550108"},
        {"first": "Kevin",   "last": "Johnson",   "email": "kevin.j@demo.com",         "phone": "+12145550109"},
        {"first": "Maria",   "last": "Garcia",    "email": "maria.garcia@demo.com",    "phone": "+12145550110"},
        {"first": "Tyler",   "last": "Brooks",    "email": "tyler.brooks@demo.com",    "phone": "+12145550111"},
        {"first": "Aisha",   "last": "Patel",     "email": "aisha.patel@demo.com",     "phone": "+12145550112"},
        {"first": "Ryan",    "last": "Kim",       "email": "ryan.kim@demo.com",        "phone": "+12145550113"},
        {"first": "Jessica", "last": "Davis",     "email": "jessica.d@demo.com",       "phone": "+12145550114"},
        {"first": "Brandon", "last": "Lee",       "email": "brandon.lee@demo.com",     "phone": "+12145550115"},
        {"first": "Natalie", "last": "Robinson",  "email": "natalie.r@demo.com",       "phone": "+12145550116"},
        {"first": "Omar",    "last": "Hassan",    "email": "omar.hassan@demo.com",     "phone": "+12145550117"},
        {"first": "Lauren",  "last": "Mitchell",  "email": "lauren.m@demo.com",        "phone": "+12145550118"},
        {"first": "Ethan",   "last": "Cooper",    "email": "ethan.c@demo.com",         "phone": "+12145550119"},
        {"first": "Rachel",  "last": "Torres",    "email": "rachel.t@demo.com",        "phone": "+12145550120"},
    ]
    DEMO_CAMPAIGNS = [
        {"name": "James Martinez",   "email": "james.martinez@demo.com",  "phone": "+12145550101", "type": "come_back",  "status": "sent",
         "msg": "Hey James! We miss you at Mario's Grill! Come back this week and enjoy 15% off your next visit. Use code: COMEBACK15"},
        {"name": "Priya Sharma",     "email": "priya.sharma@demo.com",    "phone": "+12145550102", "type": "birthday",   "status": "sent",
         "msg": "Happy Birthday Priya! From all of us at Mario's Grill, enjoy 20% off your next visit. Use code: BDAY20"},
        {"name": "Carlos Reyes",     "email": "carlos.reyes@demo.com",    "phone": "+12145550103", "type": "weekend",    "status": "sent",
         "msg": "Carlos, this weekend only — Mario's Grill is running an exclusive special for our regulars! Come in Saturday or Sunday."},
        {"name": "Ashley Thompson",  "email": "ashley.t@demo.com",        "phone": "+12145550104", "type": "loyalty",    "status": "sent",
         "msg": "Ashley, you're one of our most valued customers at Mario's Grill! Your exclusive loyalty reward is waiting — ask us when you visit!"},
        {"name": "Michael Chen",     "email": "michael.chen@demo.com",    "phone": "+12145550105", "type": "lunch",      "status": "sent",
         "msg": "Michael, join us for lunch at Mario's Grill! Fresh daily specials and your favorite dishes ready to go. Come see us!"},
        {"name": "Fatima Al-Hassan", "email": "fatima.h@demo.com",        "phone": "+12145550106", "type": "new_item",   "status": "sent",
         "msg": "Fatima, exciting news! Mario's Grill just launched something new and we think you're going to love it. Come in and be first to try it!"},
        {"name": "David Williams",   "email": "david.w@demo.com",         "phone": "+12145550107", "type": "happy_hour", "status": "sent",
         "msg": "David! Happy Hour at Mario's Grill this week — amazing drinks, great bites, unbeatable prices. Bring a friend!"},
        {"name": "Sofia Nguyen",     "email": "sofia.nguyen@demo.com",    "phone": "+12145550108", "type": "dinner",     "status": "sent",
         "msg": "Sofia, treat yourself tonight! Mario's Grill has a special dinner offer just for you — exceptional food and a memorable evening."},
        {"name": "Kevin Johnson",    "email": "kevin.j@demo.com",         "phone": "+12145550109", "type": "promotion",  "status": "sent",
         "msg": "Kevin, Mario's Grill has an exclusive promotion this week for VIP customers! Visit us and mention this message at the door."},
        {"name": "Maria Garcia",     "email": "maria.garcia@demo.com",    "phone": "+12145550110", "type": "come_back",  "status": "draft",
         "msg": "Maria, it's been a while and we miss you! Come back to Mario's Grill this week and we'll treat you like a VIP."},
        {"name": "Tyler Brooks",     "email": "tyler.brooks@demo.com",    "phone": "+12145550111", "type": "weekend",    "status": "draft",
         "msg": "Tyler! Big weekend at Mario's Grill — something special planned and we want YOU there. Saturday and Sunday only!"},
    ]

    for c in DEMO_CUSTOMERS:
        db.session.add(Customer(
            business_id=demo.id, first_name=c["first"], last_name=c["last"],
            email=c["email"], phone=c["phone"]
        ))
    for c in DEMO_CAMPAIGNS:
        db.session.add(Campaign(
            business_id=demo.id, customer_name=c["name"], customer_email=c["email"],
            customer_phone=c["phone"], campaign_type=c["type"],
            message=c["msg"], status=c["status"]
        ))
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return f"Demo setup error: {e}", 500

    session["user_id"] = demo.id
    session["is_demo"] = True
    session.permanent = True
    return redirect("/dashboard")


@app.route("/journeys")
def journeys():
    b = current_business()
    if not b:
        return redirect("/login")
    all_journeys = Journey.query.filter_by(business_id=b.id).order_by(Journey.created_at.desc()).all()
    journey_stats = {}
    for j in all_journeys:
        enrolled = CustomerJourney.query.filter_by(journey_id=j.id).count()
        completed = CustomerJourney.query.filter_by(journey_id=j.id, completed=True).count()
        active_count = CustomerJourney.query.filter_by(journey_id=j.id, completed=False, active=True).count()
        steps = JourneyStep.query.filter_by(journey_id=j.id).count()
        journey_stats[j.id] = {"enrolled": enrolled, "completed": completed, "active": active_count, "steps": steps}
    return render_template("journeys.html", journeys=all_journeys, journey_stats=journey_stats, campaign_types=get_campaign_types())


@app.route("/journeys/create", methods=["GET", "POST"])
def create_journey():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        trigger = request.form.get("trigger", "signup")
        if not name:
            flash("Journey name is required.", "error")
            return redirect("/journeys/create")
        allow_reentry = request.form.get("allow_reentry") == "1"
        freq_cap = int(request.form.get("freq_cap_per_week") or 0)
        goal = request.form.get("goal", "none")
        journey = Journey(business_id=b.id, name=name, description=description, trigger=trigger, active=True,
                          allow_reentry=allow_reentry, freq_cap_per_week=freq_cap, goal=goal)
        db.session.add(journey)
        db.session.flush()
        # Save steps
        step_orders = request.form.getlist("step_order[]")
        step_delays = request.form.getlist("step_delay[]")
        step_types = request.form.getlist("step_type[]")
        step_msgs = request.form.getlist("step_msg[]")
        step_use_ai = request.form.getlist("step_use_ai[]")
        step_channels = request.form.getlist("step_channel[]")
        step_conditions = request.form.getlist("step_condition[]")
        step_sms_fallback = request.form.getlist("step_sms_fallback[]")  # contains step index if checked
        step_sms_fallback_days = request.form.getlist("step_sms_fallback_days[]")
        for i in range(len(step_delays)):
            step = JourneyStep(
                journey_id=journey.id,
                step_order=i,
                delay_days=int(step_delays[i] or 0),
                message_type=step_types[i] if i < len(step_types) else "promotion",
                message_text=step_msgs[i] if i < len(step_msgs) else "",
                use_ai=(str(i) in step_use_ai),
                channel=step_channels[i] if i < len(step_channels) else "email",
                condition=step_conditions[i] if i < len(step_conditions) else "none",
                sms_fallback=(str(i) in step_sms_fallback),
                sms_fallback_days=int(step_sms_fallback_days[i] or 2) if i < len(step_sms_fallback_days) else 2,
            )
            db.session.add(step)
        try:
            db.session.commit()
            flash(f"Journey '{name}' created and activated!", "success")
            return redirect("/journeys")
        except Exception as e:
            db.session.rollback()
            flash(f"Error creating journey: {e}", "error")
    return render_template("create_journey.html", campaign_types=get_campaign_types())


@app.route("/journeys/<int:journey_id>/toggle", methods=["POST"])
def toggle_journey(journey_id):
    b = current_business()
    if not b:
        return redirect("/login")
    journey = Journey.query.get(journey_id)
    if journey and journey.business_id == b.id:
        journey.active = not journey.active
        db.session.commit()
        flash(f"Journey {'activated' if journey.active else 'paused'}.", "success")
    return redirect("/journeys")


@app.route("/journeys/<int:journey_id>/delete", methods=["POST"])
def delete_journey(journey_id):
    b = current_business()
    if not b:
        return redirect("/login")
    journey = Journey.query.get(journey_id)
    if journey and journey.business_id == b.id:
        JourneyStep.query.filter_by(journey_id=journey_id).delete()
        CustomerJourney.query.filter_by(journey_id=journey_id).delete()
        db.session.delete(journey)
        db.session.commit()
        flash("Journey deleted.", "success")
    return redirect("/journeys")


# ── JOURNEY ANALYTICS ───────────────────────────────────────────────────────
@app.route("/journeys/<int:journey_id>/analytics")
def journey_analytics(journey_id):
    b = current_business()
    if not b:
        return redirect("/login")
    journey = Journey.query.filter_by(id=journey_id, business_id=b.id).first_or_404()
    steps = JourneyStep.query.filter_by(journey_id=journey.id).order_by(JourneyStep.step_order).all()
    all_cjs = CustomerJourney.query.filter_by(journey_id=journey.id, business_id=b.id).all()
    total_enrolled = len(all_cjs)
    total_completed = sum(1 for cj in all_cjs if cj.completed)
    total_active = sum(1 for cj in all_cjs if not cj.completed and cj.active)
    # Per-step counts: how many customers are AT or PAST each step
    step_stats = []
    for step in steps:
        reached = sum(1 for cj in all_cjs if cj.next_step_order > step.step_order or cj.completed)
        step_stats.append({"step": step, "reached": reached,
                           "pct": round(reached / total_enrolled * 100) if total_enrolled else 0})
    # Recent enrollees
    recent = sorted(all_cjs, key=lambda c: c.enrolled_at, reverse=True)[:10]
    recent_data = []
    for cj in recent:
        customer = Customer.query.get(cj.customer_id)
        if customer:
            recent_data.append({"customer": customer, "cj": cj,
                                 "steps_total": len(steps),
                                 "progress": min(cj.next_step_order, len(steps))})
    return render_template("journey_analytics.html", journey=journey, steps=steps,
                           total_enrolled=total_enrolled, total_completed=total_completed,
                           total_active=total_active, step_stats=step_stats, recent=recent_data)


# ── CUSTOMER PROFILE (360° view like AJO Profile) ──────────────────────────
@app.route("/customers/<int:customer_id>/profile")
def customer_profile(customer_id):
    b = current_business()
    if not b:
        return redirect("/login")
    customer = Customer.query.filter_by(id=customer_id, business_id=b.id).first_or_404()
    # Campaign history for this customer
    campaigns = Campaign.query.filter_by(
        business_id=b.id, customer_email=customer.email
    ).order_by(Campaign.created_at.desc()).limit(50).all() if customer.email else []
    # Journey memberships
    journey_memberships = []
    cjs = CustomerJourney.query.filter_by(customer_id=customer.id, business_id=b.id).all()
    for cj in cjs:
        journey = Journey.query.get(cj.journey_id)
        if journey:
            steps_total = JourneyStep.query.filter_by(journey_id=journey.id).count()
            journey_memberships.append({
                "journey": journey,
                "cj": cj,
                "steps_total": steps_total,
                "progress": min(cj.next_step_order, steps_total),
            })
    return render_template("customer_profile.html", customer=customer, business=b,
                           campaigns=campaigns, journey_memberships=journey_memberships)


# ── AUDIENCES (like AJO Audiences / Segments) ───────────────────────────────
def compute_audience_members(audience, business_id):
    """Evaluate JSON rules against all customers and return matching list."""
    import json as _json
    try:
        rules = _json.loads(audience.rules or "[]")
    except Exception:
        rules = []
    customers = Customer.query.filter_by(business_id=business_id, unsubscribed=False).all()
    if not rules:
        return customers
    now = datetime.utcnow()
    matched = []
    for c in customers:
        ok = True
        for rule in rules:
            field = rule.get("field", "")
            op = rule.get("op", "")
            val = rule.get("value", "")
            try:
                if field == "visit_count":
                    cv = c.visit_count or 0
                    val = int(val)
                    if op == "gte" and not (cv >= val): ok = False
                    elif op == "lte" and not (cv <= val): ok = False
                    elif op == "eq" and not (cv == val): ok = False
                elif field == "loyalty_points":
                    cv = c.loyalty_points or 0
                    val = int(val)
                    if op == "gte" and not (cv >= val): ok = False
                    elif op == "lte" and not (cv <= val): ok = False
                elif field == "days_since_visit":
                    ref = c.last_visit or c.created_at
                    days = (now - ref).days
                    val = int(val)
                    if op == "gte" and not (days >= val): ok = False
                    elif op == "lte" and not (days <= val): ok = False
                elif field == "days_since_signup":
                    days = (now - c.created_at).days
                    val = int(val)
                    if op == "gte" and not (days >= val): ok = False
                    elif op == "lte" and not (days <= val): ok = False
                elif field == "tag":
                    tags = [t.strip().lower() for t in (c.tags or "").split(",")]
                    if op == "contains" and val.lower() not in tags: ok = False
                    elif op == "not_contains" and val.lower() in tags: ok = False
                elif field == "birthday_this_month":
                    if c.dob:
                        try:
                            month = int(c.dob.split("-")[1]) if "-" in c.dob else int(c.dob.split("/")[1])
                            if month != now.month: ok = False
                        except Exception: ok = False
                    else:
                        ok = False
                elif field == "has_email":
                    if op == "eq" and val == "true" and not c.email: ok = False
                    elif op == "eq" and val == "false" and c.email: ok = False
                elif field == "has_phone":
                    if op == "eq" and val == "true" and not c.phone: ok = False
                    elif op == "eq" and val == "false" and c.phone: ok = False
            except Exception:
                pass
            if not ok:
                break
        if ok:
            matched.append(c)
    return matched


@app.route("/audiences")
def audiences():
    b = current_business()
    if not b:
        return redirect("/login")
    all_audiences = Audience.query.filter_by(business_id=b.id).order_by(Audience.created_at.desc()).all()
    audience_sizes = {}
    for a in all_audiences:
        audience_sizes[a.id] = len(compute_audience_members(a, b.id))
    all_journeys = Journey.query.filter_by(business_id=b.id, active=True).all()
    return render_template("audiences.html", audiences=all_audiences,
                           audience_sizes=audience_sizes, journeys=all_journeys)


@app.route("/audiences/create", methods=["GET", "POST"])
def create_audience():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        import json as _json
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        fields = request.form.getlist("rule_field[]")
        ops = request.form.getlist("rule_op[]")
        vals = request.form.getlist("rule_value[]")
        rules = []
        for f, o, v in zip(fields, ops, vals):
            if f and o:
                rules.append({"field": f, "op": o, "value": v})
        if not name:
            flash("Audience name is required.", "error")
        else:
            aud = Audience(
                business_id=b.id,
                name=name,
                description=description,
                rules=_json.dumps(rules),
            )
            db.session.add(aud)
            db.session.commit()
            flash(f"Audience '{name}' created!", "success")
            return redirect("/audiences")
    return render_template("create_audience.html")


@app.route("/audiences/<int:audience_id>/delete", methods=["POST"])
def delete_audience(audience_id):
    b = current_business()
    if not b:
        return redirect("/login")
    aud = Audience.query.filter_by(id=audience_id, business_id=b.id).first()
    if aud:
        db.session.delete(aud)
        db.session.commit()
        flash("Audience deleted.", "success")
    return redirect("/audiences")


@app.route("/audiences/<int:audience_id>/enroll", methods=["POST"])
def enroll_audience(audience_id):
    """Enroll all audience members into a selected journey."""
    b = current_business()
    if not b:
        return redirect("/login")
    aud = Audience.query.filter_by(id=audience_id, business_id=b.id).first_or_404()
    journey_id = request.form.get("journey_id", type=int)
    if not journey_id:
        flash("Please select a journey.", "error")
        return redirect("/audiences")
    journey = Journey.query.filter_by(id=journey_id, business_id=b.id).first_or_404()
    members = compute_audience_members(aud, b.id)
    enrolled = 0
    now = datetime.utcnow()
    for customer in members:
        existing = CustomerJourney.query.filter_by(journey_id=journey.id, customer_id=customer.id).first()
        if existing:
            continue
        first_step = JourneyStep.query.filter_by(journey_id=journey.id, step_order=0).first()
        delay = first_step.delay_days if first_step else 0
        cj = CustomerJourney(
            journey_id=journey.id,
            customer_id=customer.id,
            business_id=b.id,
            next_step_order=0,
            enrolled_at=now,
            next_step_at=now + timedelta(days=delay),
        )
        db.session.add(cj)
        enrolled += 1
    db.session.commit()
    flash(f"Enrolled {enrolled} customers from '{aud.name}' into journey '{journey.name}'.", "success")
    return redirect("/audiences")


@app.route("/audiences/<int:audience_id>/preview")
def preview_audience(audience_id):
    """Return JSON list of matching customer names/emails for live preview."""
    b = current_business()
    if not b:
        return {"error": "unauthorized"}, 401
    aud = Audience.query.filter_by(id=audience_id, business_id=b.id).first_or_404()
    members = compute_audience_members(aud, b.id)
    return {"count": len(members), "members": [
        {"name": f"{c.first_name} {c.last_name or ''}".strip(), "email": c.email or "", "phone": c.phone or ""}
        for c in members[:20]
    ]}


@app.route("/audiences/preview-rules", methods=["POST"])
def preview_audience_rules():
    """Live-preview audience size as user builds rules."""
    import json as _json
    b = current_business()
    if not b:
        return {"error": "unauthorized"}, 401
    data = request.get_json(silent=True) or {}
    rules_raw = data.get("rules", [])
    fake_aud = type("A", (), {"rules": _json.dumps(rules_raw)})()
    members = compute_audience_members(fake_aud, b.id)
    return {"count": len(members)}


@app.route("/analytics")
def analytics():
    b = current_business()
    if not b:
        return redirect("/login")

    from datetime import timedelta

    all_campaigns = Campaign.query.filter_by(business_id=b.id).all()
    sent = [c for c in all_campaigns if c.status == "sent"]

    total_sent = len(sent)
    total_opens = sum(c.open_count or 0 for c in sent)
    total_clicks = sum(c.click_count or 0 for c in sent)
    open_rate = round((total_opens / total_sent * 100), 1) if total_sent else 0
    click_rate = round((total_clicks / total_sent * 100), 1) if total_sent else 0

    # Per campaign type breakdown
    type_stats = {}
    for c in sent:
        t = c.campaign_type
        if t not in type_stats:
            type_stats[t] = {"sent": 0, "opens": 0, "clicks": 0}
        type_stats[t]["sent"] += 1
        type_stats[t]["opens"] += c.open_count or 0
        type_stats[t]["clicks"] += c.click_count or 0
    for t in type_stats:
        s = type_stats[t]["sent"]
        type_stats[t]["open_rate"] = round(type_stats[t]["opens"] / s * 100, 1) if s else 0
        type_stats[t]["click_rate"] = round(type_stats[t]["clicks"] / s * 100, 1) if s else 0
        type_stats[t]["label"] = get_campaign_types().get(t, t)

    # Last 30 days daily sends
    today = datetime.utcnow().date()
    days_30 = [(today - timedelta(days=i)) for i in range(29, -1, -1)]
    daily_map = {}
    for c in sent:
        d = c.created_at.date()
        daily_map[d] = daily_map.get(d, 0) + 1
    daily_labels = [d.strftime("%b %d") for d in days_30]
    daily_data = [daily_map.get(d, 0) for d in days_30]

    # Top 5 customers by opens + clicks
    customer_engagement = {}
    for c in sent:
        key = (c.customer_name, c.customer_email)
        if key not in customer_engagement:
            customer_engagement[key] = {"opens": 0, "clicks": 0, "campaigns": 0}
        customer_engagement[key]["opens"] += c.open_count or 0
        customer_engagement[key]["clicks"] += c.click_count or 0
        customer_engagement[key]["campaigns"] += 1
    top_customers = sorted(
        [{"name": k[0], "email": k[1], **v} for k, v in customer_engagement.items()],
        key=lambda x: x["opens"] + x["clicks"], reverse=True
    )[:5]

    return render_template("analytics.html",
        total_sent=total_sent,
        total_opens=total_opens,
        total_clicks=total_clicks,
        open_rate=open_rate,
        click_rate=click_rate,
        type_stats=type_stats,
        daily_labels=daily_labels,
        daily_data=daily_data,
        top_customers=top_customers,
        campaign_types=get_campaign_types(),
    )

@app.route("/settings", methods=["GET", "POST"])
def settings():
    b = current_business()
    if not b:
        return redirect("/login")
    if request.method == "POST":
        b.business_name = request.form.get("business_name", b.business_name).strip() or b.business_name
        b.owner_name = request.form.get("owner_name", b.owner_name).strip() or b.owner_name
        b.address = request.form.get("address", "").strip()
        b.phone = request.form.get("phone", "").strip()
        b.website = request.form.get("website", "").strip()
        try:
            db.session.commit()
            flash("Settings saved!", "success")
        except:
            db.session.rollback()
            flash("Error saving settings.", "error")
        return redirect("/settings")
    return render_template("settings.html", business=b)

@app.route("/upgrade")
def upgrade():
    b = current_business()
    if not b:
        return redirect("/login")
    return render_template("upgrade.html", plan=b.plan, business_name=b.business_name)

@app.route("/upgrade/starter", methods=["POST"])
def upgrade_starter():
    b = current_business()
    if not b:
        return redirect("/login")
    try:
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            mode="subscription",
            line_items=[{"price": STRIPE_PRICE_STARTER, "quantity": 1}],
            success_url=url_for("upgrade_success", _external=True) + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=url_for("upgrade", _external=True),
            customer_email=b.email
        )
        return redirect(checkout_session.url)
    except Exception as e:
        flash(f"Error: {e}", "error")
        return redirect("/upgrade")

@app.route("/upgrade/pro", methods=["POST"])
def upgrade_pro():
    b = current_business()
    if not b:
        return redirect("/login")
    try:
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            mode="subscription",
            line_items=[{"price": STRIPE_PRICE_PRO, "quantity": 1}],
            success_url=url_for("upgrade_success", _external=True) + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=url_for("upgrade", _external=True),
            customer_email=b.email
        )
        return redirect(checkout_session.url)
    except Exception as e:
        flash(f"Error: {e}", "error")
        return redirect("/upgrade")

@app.route("/upgrade/success")
def upgrade_success():
    b = current_business()
    if not b:
        return redirect("/login")
    session_id = request.args.get("session_id", "")
    if not session_id:
        flash("Invalid upgrade link.", "error")
        return redirect("/upgrade")
    try:
        checkout_session = stripe.checkout.Session.retrieve(session_id)
        if checkout_session.payment_status not in ("paid", "no_payment_required"):
            flash("Payment not completed. Please try again.", "error")
            return redirect("/upgrade")
        price_id = checkout_session.line_items.data[0].price.id if checkout_session.line_items else None
        if price_id == STRIPE_PRICE_PRO:
            plan = "pro"
        elif price_id == STRIPE_PRICE_STARTER:
            plan = "starter"
        else:
            # Fallback: derive from amount
            plan = "pro" if checkout_session.amount_total and checkout_session.amount_total > 1900 else "starter"
    except Exception as e:
        print(f"Stripe session verify error: {e}")
        flash("Could not verify payment. Contact support.", "error")
        return redirect("/upgrade")
    if b:
        b.plan = plan
        db.session.commit()
    flash(f"You're now on the {plan.title()} plan!", "success")
    return redirect("/dashboard")

@app.errorhandler(403)
def forbidden(e):
    return render_template("404.html"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500

# ============================================
# ADMIN PORTAL
# ============================================

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

def admin_required():
    return session.get("is_admin") is True

@app.route("/admin")
def admin_index():
    return redirect("/admin/login")

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if admin_required():
        return redirect("/admin/dashboard")
    if request.method == "POST":
        pw = request.form.get("password", "")
        if ADMIN_PASSWORD and pw == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect("/admin/dashboard")
        flash("Invalid admin password.", "error")
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect("/admin/login")

@app.route("/admin/dashboard")
def admin_dashboard():
    if not admin_required():
        return redirect("/admin/login")
    businesses = Business.query.order_by(Business.created_at.desc()).all()
    stats = []
    for b in businesses:
        customer_count = Customer.query.filter_by(business_id=b.id).count()
        campaign_count = Campaign.query.filter_by(business_id=b.id).count()
        sent_count = Campaign.query.filter_by(business_id=b.id, status="sent").count()
        total_opens = db.session.query(db.func.sum(Campaign.open_count)).filter_by(business_id=b.id).scalar() or 0
        stats.append({
            "business": b,
            "customers": customer_count,
            "campaigns": campaign_count,
            "sent": sent_count,
            "opens": total_opens,
        })
    total_businesses = len(businesses)
    total_customers = Customer.query.count()
    total_sent = Campaign.query.filter_by(status="sent").count()
    total_opens = db.session.query(db.func.sum(Campaign.open_count)).scalar() or 0
    plan_counts = {
        "free": sum(1 for b in businesses if b.plan == "free"),
        "starter": sum(1 for b in businesses if b.plan == "starter"),
        "pro": sum(1 for b in businesses if b.plan == "pro"),
    }
    mrr = plan_counts["starter"] * 19 + plan_counts["pro"] * 50
    all_campaign_types = CampaignTypeModel.query.order_by(CampaignTypeModel.sort_order, CampaignTypeModel.label).all()
    return render_template("admin_dashboard.html",
        stats=stats,
        total_businesses=total_businesses,
        total_customers=total_customers,
        total_sent=total_sent,
        total_opens=total_opens,
        plan_counts=plan_counts,
        mrr=mrr,
        campaign_types=all_campaign_types,
    )

@app.route("/admin/add-business", methods=["POST"])
def admin_add_business():
    if not admin_required():
        return redirect("/admin/login")
    business_name = request.form.get("business_name", "").strip()
    owner_name = request.form.get("owner_name", "").strip()
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "").strip()
    plan = request.form.get("plan", "free")
    phone = request.form.get("phone", "").strip()
    address = request.form.get("address", "").strip()
    if not all([business_name, owner_name, email, password]):
        flash("Business name, owner name, email, and password are required.", "error")
        return redirect("/admin/dashboard")
    if Business.query.filter_by(email=email).first():
        flash(f"Email {email} is already registered.", "error")
        return redirect("/admin/dashboard")
    hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    b = Business(
        business_name=business_name,
        owner_name=owner_name,
        email=email,
        password=hashed,
        plan=plan,
        phone=phone or None,
        address=address or None,
    )
    try:
        db.session.add(b)
        db.session.commit()
        flash(f"Business '{business_name}' created successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error creating business: {e}", "error")
    return redirect("/admin/dashboard")

@app.route("/admin/change-plan/<int:business_id>", methods=["POST"])
def admin_change_plan(business_id):
    if not admin_required():
        return redirect("/admin/login")
    b = Business.query.get(business_id)
    if b:
        b.plan = request.form.get("plan", b.plan)
        db.session.commit()
    return redirect("/admin/dashboard")

@app.route("/admin/login-as/<int:business_id>", methods=["POST"])
def admin_login_as(business_id):
    if not admin_required():
        return redirect("/admin/login")
    b = Business.query.get(business_id)
    if b:
        session["user_id"] = b.id
        session["admin_impersonating"] = True
    return redirect("/dashboard")

@app.route("/admin/delete-business/<int:business_id>", methods=["POST"])
def admin_delete_business(business_id):
    if not admin_required():
        return redirect("/admin/login")
    b = Business.query.get(business_id)
    if b:
        Campaign.query.filter_by(business_id=b.id).delete()
        Customer.query.filter_by(business_id=b.id).delete()
        AutomationRule.query.filter_by(business_id=b.id).delete()
        Promotion.query.filter_by(business_id=b.id).delete()
        db.session.delete(b)
        db.session.commit()
    return redirect("/admin/dashboard")

@app.route("/admin/campaign-types/add", methods=["POST"])
def admin_add_campaign_type():
    if not admin_required():
        return redirect("/admin/login")
    key = request.form.get("key", "").strip().lower().replace(" ", "_")
    label = request.form.get("label", "").strip()
    if not key or not label:
        flash("Key and label are required.", "error")
        return redirect("/admin/dashboard")
    if CampaignTypeModel.query.filter_by(key=key).first():
        flash(f"Key '{key}' already exists.", "error")
        return redirect("/admin/dashboard")
    max_order = db.session.query(db.func.max(CampaignTypeModel.sort_order)).scalar() or 0
    db.session.add(CampaignTypeModel(key=key, label=label, sort_order=max_order + 1))
    db.session.commit()
    flash(f"Campaign type '{label}' added.", "success")
    return redirect("/admin/dashboard")

@app.route("/admin/campaign-types/edit/<int:type_id>", methods=["POST"])
def admin_edit_campaign_type(type_id):
    if not admin_required():
        return redirect("/admin/login")
    ct = CampaignTypeModel.query.get(type_id)
    if not ct:
        flash("Not found.", "error")
        return redirect("/admin/dashboard")
    ct.label = request.form.get("label", ct.label).strip() or ct.label
    ct.active = request.form.get("active") == "on"
    db.session.commit()
    flash("Updated.", "success")
    return redirect("/admin/dashboard")

@app.route("/admin/campaign-types/delete/<int:type_id>", methods=["POST"])
def admin_delete_campaign_type(type_id):
    if not admin_required():
        return redirect("/admin/login")
    ct = CampaignTypeModel.query.get(type_id)
    if ct:
        db.session.delete(ct)
        db.session.commit()
        flash(f"Deleted '{ct.label}'.", "success")
    return redirect("/admin/dashboard")

# ============================================
# ONBOARDING WIZARD  /onboarding
# ============================================

@app.route("/onboarding", methods=["GET", "POST"])
def onboarding():
    b = current_business()
    if not b:
        return redirect("/login")

    _ensure_slug(b)
    profile = BusinessProfile.query.filter_by(business_id=b.id).first()
    if not profile:
        profile = BusinessProfile(business_id=b.id)
        db.session.add(profile)
        db.session.commit()

    # Already set up — go to dashboard
    if profile.setup_complete and request.method == "GET" and not request.args.get("redo"):
        return redirect("/dashboard")

    if request.method == "POST":
        # Save all onboarding fields
        new_name = request.form.get("business_name", "").strip()
        if new_name:
            b.business_name = new_name
        b.phone   = request.form.get("phone", b.phone or "").strip() or b.phone
        b.address = request.form.get("address", b.address or "").strip() or b.address
        b.website = request.form.get("website", b.website or "").strip() or b.website

        profile.cuisine_type    = request.form.get("cuisine_type", "").strip()
        profile.signature_dish  = request.form.get("signature_dish", "").strip()
        profile.special_offer   = request.form.get("special_offer", "").strip()
        profile.slow_days       = request.form.get("slow_days", "").strip()
        profile.tone            = request.form.get("tone", "friendly")
        profile.auto_welcome    = True
        profile.auto_birthday   = True
        profile.auto_winback    = True
        profile.auto_review     = True
        profile.auto_loyalty    = True
        profile.auto_weekly     = request.form.get("auto_weekly") == "on"
        profile.auto_flash      = bool(profile.slow_days)
        profile.setup_complete  = True

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving: {e}", "error")
            return redirect("/onboarding")

        join_url = request.host_url.rstrip("/") + f"/join/{b.slug}"
        return render_template("onboarding_done.html", business=b, join_url=join_url)

    return render_template("onboarding.html", business=b, profile=profile)


# ============================================
# AUTOPILOT TEST SEND  /autopilot/test
# ============================================

@app.route("/autopilot/test", methods=["POST"])
def autopilot_test():
    """Send a sample of each active automation to the business owner's own email."""
    b = current_business()
    if not b:
        return redirect("/login")
    if b.plan == "free":
        flash("Auto-pilot requires Starter or Pro plan.", "error")
        return redirect("/upgrade")

    profile = BusinessProfile.query.filter_by(business_id=b.id).first()
    offer = (profile.special_offer if profile and profile.special_offer else "a special offer")
    dish  = (profile.signature_dish if profile and profile.signature_dish else "our signature dish")
    tone_hint = f" Use a {profile.tone} tone." if profile and profile.tone else ""

    results = []

    def _test_send(label, campaign_type, prompt_extra=""):
        prompt = (
            f"Write a short marketing SMS/email for {b.owner_name} from {b.business_name}. "
            f"Campaign type: {label}. Mention: {offer}.{tone_hint} {prompt_extra} "
            f"Keep it under 3 sentences, no markdown, include 1-2 emojis."
        )
        try:
            msg = clean_ai_text(_call_openai(prompt)) if client else f"Hi {b.owner_name}! Test message for '{label}' from {b.business_name}. {offer}."
        except Exception:
            msg = f"Hi {b.owner_name}! Test message for '{label}' from {b.business_name}. {offer}."

        # Send to business owner email
        ok = send_email(
            b.email,
            f"[TEST] {label} — {b.business_name}",
            msg,
            customer_name=b.owner_name,
            business_name=b.business_name,
            campaign_type=campaign_type,
            business_address=b.address or "",
            business_phone=b.phone or "",
            business_website=b.website or "",
            business_reply_email=b.email,
        )
        results.append({"label": label, "msg": msg, "ok": ok})

    if not profile or not profile.setup_complete:
        flash("Please complete your AI Setup first before testing.", "error")
        return redirect("/setup")

    if profile.auto_welcome:
        _test_send("Welcome Message", "loyalty", "This is sent when a new customer signs up.")
    if profile.auto_weekly:
        _test_send("Weekly Special", "weekend", f"Feature {dish}.")
    if profile.auto_flash:
        _test_send("Flash Deal", "promotion", f"Slow day deal — urgent, limited time.")
    if profile.auto_birthday:
        _test_send("Birthday Offer", "birthday", "Include a birthday discount.")
    if profile.auto_winback:
        _test_send("Win-Back (30 days)", "come_back", "Customer hasn't visited in 30 days.")

    if not results:
        flash("No automations are turned on. Enable some in AI Setup first.", "error")
        return redirect("/autopilot")

    sent_count = sum(1 for r in results if r["ok"])
    flash(f"✅ Test sent! {sent_count}/{len(results)} automation previews sent to {b.email}. Check your inbox.", "success")
    return redirect("/autopilot")


# ============================================
# PUBLIC OPT-IN PAGE  /join/<slug>
# ============================================

def _ensure_slug(b):
    """Make sure a business has a slug; save it if missing."""
    if not b.slug:
        b.slug = make_slug(b.business_name, b.id)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

@app.route("/join/<slug>", methods=["GET", "POST"])
def customer_optin(slug):
    b = Business.query.filter_by(slug=slug).first()
    if not b:
        return render_template("optin_notfound.html"), 404

    if request.method == "POST":
        first_name = request.form.get("first_name", "").strip()
        last_name  = request.form.get("last_name", "").strip()
        email      = request.form.get("email", "").strip()
        phone      = request.form.get("phone", "").strip()
        dob        = request.form.get("dob", "").strip()
        sms_consent = request.form.get("sms_consent") == "on"

        if not first_name:
            flash("First name is required.", "error")
            return redirect(f"/join/{slug}")
        if not email and not phone:
            flash("Please provide at least an email or phone number.", "error")
            return redirect(f"/join/{slug}")
        if phone and not sms_consent:
            flash("Please check the SMS consent box to receive text messages.", "error")
            return redirect(f"/join/{slug}")

        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

        # Check for duplicate by phone or email within this business
        existing = None
        if email:
            existing = Customer.query.filter_by(business_id=b.id, email=email).first()
        if not existing and phone:
            existing = Customer.query.filter_by(business_id=b.id, phone=phone).first()

        if existing:
            # Update opt-in status if they re-submit
            if phone and sms_consent:
                existing.phone = phone
                existing.sms_opted_in = True
                existing.sms_opted_in_at = datetime.now(timezone.utc)
                existing.sms_opt_in_ip = ip
            db.session.commit()
            return render_template("optin_success.html", business=b, already=True)

        customer = Customer(
            business_id=b.id,
            first_name=first_name,
            last_name=last_name or "",
            email=email or None,
            phone=phone or None,
            dob=dob or None,
            sms_opted_in=sms_consent and bool(phone),
            sms_opted_in_at=datetime.now(timezone.utc) if (sms_consent and phone) else None,
            sms_opt_in_ip=ip if (sms_consent and phone) else None,
        )
        db.session.add(customer)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("Something went wrong. Please try again.", "error")
            return redirect(f"/join/{slug}")

        # Send welcome SMS if opted in and automation is on
        profile = BusinessProfile.query.filter_by(business_id=b.id).first()
        if customer.sms_opted_in and customer.phone and profile and profile.auto_welcome:
            offer = profile.special_offer or "a special welcome offer"
            welcome_msg = (
                f"Welcome to {b.business_name}, {first_name}! 🎉 "
                f"Thanks for joining — enjoy {offer} on your next visit!"
            )
            send_sms(customer.phone, welcome_msg)

        return render_template("optin_success.html", business=b, already=False)

    _ensure_slug(b)
    consent_text = (
        f"By checking this box, you agree to receive marketing text messages from "
        f"{b.business_name}. Message frequency varies. Msg & data rates may apply. "
        f"Reply STOP to unsubscribe at any time. Reply HELP for help."
    )
    return render_template("customer_optin.html", business=b, consent_text=consent_text)


# ============================================
# TWILIO WEBHOOK — handle STOP / HELP replies
# ============================================

@app.route("/customer/<int:customer_id>/visit", methods=["POST"])
def log_visit(customer_id):
    """Log a customer visit: increment visit_count, set last_visit, add loyalty points."""
    b = current_business()
    if not b:
        return jsonify({"ok": False, "error": "not logged in"}), 401
    c = Customer.query.filter_by(id=customer_id, business_id=b.id).first_or_404()
    c.visit_count    = (c.visit_count or 0) + 1
    c.last_visit     = datetime.now(timezone.utc)
    c.loyalty_points = (c.loyalty_points or 0) + 10
    try:
        db.session.commit()
        return jsonify({"ok": True, "visits": c.visit_count, "points": c.loyalty_points})
    except Exception:
        db.session.rollback()
        return jsonify({"ok": False}), 500


@app.route("/twilio/webhook", methods=["POST"])
def twilio_webhook():
    """Twilio calls this when a customer replies STOP, HELP, etc."""
    from_number = request.form.get("From", "").strip()
    body        = request.form.get("Body", "").strip().upper()

    if from_number and body in ("STOP", "STOPALL", "UNSUBSCRIBE", "CANCEL", "QUIT"):
        # Mark all customers with this phone as unsubscribed + sms_opted_in=False
        customers = Customer.query.filter_by(phone=from_number).all()
        for c in customers:
            c.unsubscribed   = True
            c.sms_opted_in   = False
        try:
            db.session.commit()
            print(f"STOP received from {from_number} — {len(customers)} customer(s) unsubscribed.")
        except Exception:
            db.session.rollback()

    # Twilio expects a TwiML response (can be empty)
    return '<?xml version="1.0" encoding="UTF-8"?><Response></Response>', 200, {"Content-Type": "text/xml"}


# ============================================
# BUSINESS SETUP / ONBOARDING  /setup
# ============================================

@app.route("/setup", methods=["GET", "POST"])
def business_setup():
    b = current_business()
    if not b:
        return redirect("/login")

    _ensure_slug(b)
    profile = BusinessProfile.query.filter_by(business_id=b.id).first()
    if not profile:
        profile = BusinessProfile(business_id=b.id)
        db.session.add(profile)
        db.session.commit()

    if request.method == "POST":
        profile.cuisine_type     = request.form.get("cuisine_type", "").strip()
        profile.signature_dish   = request.form.get("signature_dish", "").strip()
        profile.special_offer    = request.form.get("special_offer", "").strip()
        profile.slow_days        = request.form.get("slow_days", "").strip()
        profile.peak_days        = request.form.get("peak_days", "").strip()
        profile.tone             = request.form.get("tone", "friendly")
        profile.timezone         = request.form.get("timezone", "America/Chicago")
        profile.auto_welcome          = request.form.get("auto_welcome") == "on"
        profile.auto_weekly           = request.form.get("auto_weekly") == "on"
        profile.auto_flash            = request.form.get("auto_flash") == "on"
        profile.auto_birthday         = request.form.get("auto_birthday") == "on"
        profile.auto_winback          = request.form.get("auto_winback") == "on"
        profile.auto_review           = request.form.get("auto_review") == "on"
        profile.auto_loyalty          = request.form.get("auto_loyalty") == "on"
        profile.google_review_url     = request.form.get("google_review_url", "").strip()
        profile.loyalty_reward_visits = int(request.form.get("loyalty_reward_visits") or 5)
        profile.weekly_send_day       = request.form.get("weekly_send_day", "Tuesday")
        profile.language              = request.form.get("language", "English")
        profile.setup_complete        = True

        # Also update business website if provided
        website = request.form.get("website", "").strip()
        if website:
            b.website = website

        try:
            db.session.commit()
            flash("Setup saved! Your AI auto-pilot is active. 🚀", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving: {e}", "error")
        return redirect("/setup")

    join_url = request.host_url.rstrip("/") + f"/join/{b.slug}"
    return render_template("setup.html", business=b, profile=profile, join_url=join_url)


# ============================================
# AUTO-PILOT DASHBOARD  /autopilot
# ============================================

@app.route("/autopilot")
def autopilot():
    b = current_business()
    if not b:
        return redirect("/login")
    if b.plan == "free":
        flash("Auto-pilot requires Starter or Pro plan.", "error")
        return redirect("/upgrade")

    _ensure_slug(b)
    profile = BusinessProfile.query.filter_by(business_id=b.id).first()

    # Recent automated campaigns (last 30 days)
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    auto_types = ["welcome", "birthday", "come_back", "loyalty", "weekly_special", "flash_deal"]
    recent = Campaign.query.filter(
        Campaign.business_id == b.id,
        Campaign.status == "sent",
        Campaign.created_at >= thirty_days_ago,
        Campaign.campaign_type.in_(auto_types)
    ).order_by(Campaign.created_at.desc()).limit(50).all()

    total_auto_sent   = len(recent)
    total_auto_opens  = sum(c.open_count or 0 for c in recent)
    sms_opted_count   = Customer.query.filter_by(business_id=b.id, sms_opted_in=True, unsubscribed=False).count()
    join_url          = request.host_url.rstrip("/") + f"/join/{b.slug}"
    avg_visits        = db.session.query(db.func.avg(Customer.visit_count)).filter_by(business_id=b.id).scalar() or 0
    total_loyalty_pts = db.session.query(db.func.sum(Customer.loyalty_points)).filter_by(business_id=b.id).scalar() or 0

    return render_template(
        "autopilot.html",
        business=b,
        profile=profile,
        recent=recent,
        total_auto_sent=total_auto_sent,
        total_auto_opens=total_auto_opens,
        sms_opted_count=sms_opted_count,
        join_url=join_url,
        avg_visits=round(float(avg_visits), 1),
        total_loyalty_pts=int(total_loyalty_pts),
    )


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.getenv("DEBUG", "False").lower() == "true"
    app.run(host="127.0.0.1", port=port, debug=debug_mode, use_reloader=False)
